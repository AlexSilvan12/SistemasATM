import shutil
import tkinter as tk
from tkinter import filedialog
from ControlSolicitudes import Control_Solicitudes
from gastos_contrato import costos_contrato
from tkinter import messagebox, ttk
from PIL import Image, ImageTk
from datetime import date, datetime
from database import conectar_bd
from proveedores import cargar_proveedores
from utils import ruta_relativa, centrar_ventana, convertir_excel_a_pdf, salir
from login import usuario_actual
from openpyxl import load_workbook
import mysql.connector
import os
from tkcalendar import DateEntry
from openpyxl.styles import Alignment, Border, Side
from openpyxl.drawing.image import Image as ExcelImage

factura_seleccionada = None
CARPETA_COMPROBANTES = os.path.join(
            os.environ['USERPROFILE'], 
            'OneDrive - ATI', 
            'Documentos de AppGestor', 
            'Comprobantes de Pago'
        )

         
CARPETA_SOLICITUDES = os.path.join(
    os.environ['USERPROFILE'],
    'OneDrive - ATI',
    'Documentos de AppGestor',
    'Solicitudes de Pago'
    )

# Función para conectar con las solicitudes almacenadas en la base de datos
def cargar_solicitudes(tree):
    # Limpiar datos previos
    for row in tree.get_children():
        tree.delete(row)

    conexion = None
    cursor = None

    try:
        conexion = conectar_bd()
        if conexion is None:
            print("❌ No se pudo establecer la conexión")
            return

        cursor = conexion.cursor()
        cursor.execute("""
            SELECT id_solicitud, fecha_solicitud, importe, fecha_limite_pago, estado, archivo 
            FROM SolicitudesPago 
            WHERE estado = 'Autorizado' OR estado = 'Pendiente'
            ORDER BY id_solicitud ASC
        """)

        # Configurar tags para colores
        tree.tag_configure("autorizado", background="#90EE90")  # Verde claro
        tree.tag_configure("pendiente", background="#FFD580")   # Amarillo claro
        tree.tag_configure("rechazado", background="#FF7F7F")   # Rojo claro (por si después lo usas)

        for solicitud in cursor.fetchall():
            estado = solicitud[4]  # La columna 'estado' está en índice 4

            if estado.lower() == "autorizado":
                tree.insert("", "end", values=solicitud, tags=("autorizado",))
            elif estado.lower() == "pendiente":
                tree.insert("", "end", values=solicitud, tags=("pendiente",))
            elif estado.lower() == "rechazado":
                tree.insert("", "end", values=solicitud, tags=("rechazado",))
            else:
                tree.insert("", "end", values=solicitud)

    except Exception as e:
        print(f"❌ Error al cargar solicitudes de pago: {e}")

    finally:
        if cursor is not None:
            cursor.close()
        if conexion is not None:
            conexion.close()

def cargar_autorizaciones(tree):
    tree.delete(*tree.get_children())  # Limpiar tabla

    try:
        conexion = conectar_bd()
        cursor = conexion.cursor()
        cursor.execute("""
    SELECT id_autorizacion, fecha_solicitud, monto, fecha_limite_pago 
    FROM autorizacionescompra
    WHERE id_autorizacion NOT IN (
        SELECT id_autorizacion FROM SolicitudesPago
    )
    """)
        for row in cursor.fetchall():
            tree.insert("", tk.END, values=row)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo cargar autorizaciones: {e}")
    finally:
        if cursor: cursor.close()
        if conexion: conexion.close()


def marcar_como_pagado(tree, usuario_actual, ventana_padre):
    if usuario_actual["rol"] != "Contador":
        messagebox.showwarning("Acceso denegado", "No tiene permisos para realizar esta acción.", parent=ventana_padre)
        return

    seleccion = tree.focus()
    if not seleccion:
        messagebox.showwarning("Sin selección", "Selecciona una solicitud para marcar como pagada.", parent=ventana_padre)
        return

    valores = tree.item(seleccion, "values")
    if not valores:
        return

    id_solicitud = valores[0]

    conexion = None
    cursor = None
    try:
        conexion = conectar_bd()
        cursor = conexion.cursor()

        cursor.execute("SELECT estado FROM solicitudespago WHERE id_solicitud = %s", (id_solicitud,))
        estado_actual = cursor.fetchone()
        if not estado_actual or estado_actual[0] != "Autorizado":
            messagebox.showinfo("Información", "La Solicitud de Pago aun no ha sido autorizada.", parent=ventana_padre)
            return

        # Seleccionar comprobante
        ruta_archivo = filedialog.askopenfilename(
            title="Seleccionar comprobante de pago",
            filetypes=[("Archivos PDF", "*.pdf"), ("Imágenes", "*.jpg;*.jpeg;*.png"), ("Todos los archivos", "*.*")],
            parent=ventana_padre
        )
        if not ruta_archivo:
            messagebox.showwarning("Cancelado", "No seleccionaste ningún comprobante.", parent=ventana_padre)
            return
        
        # Asegurarnos de que la carpeta base exista
        if not os.path.exists(CARPETA_COMPROBANTES):
            os.makedirs(CARPETA_COMPROBANTES)

        extension = os.path.splitext(ruta_archivo)[1]
        nombre_final = f"Comprobante_{id_solicitud}{extension}"  # Solo el nombre
        ruta_destino = os.path.join(CARPETA_COMPROBANTES, nombre_final)

        hoy = date.today()

        # Guardar solo el nombre del archivo en la BD
        cursor.execute("""
            UPDATE solicitudespago 
            SET estado = 'Pagado', fecha_pago = %s, comprobante = %s
            WHERE id_solicitud = %s
        """, (hoy, nombre_final, id_solicitud))  # << Solo nombre_final, NO la ruta completa

        if cursor.rowcount == 0:
            messagebox.showerror("Error", "No se pudo actualizar la solicitud en la base de datos.", parent=ventana_padre)
            conexion.rollback()
            return

        try:
            shutil.copy2(ruta_archivo, ruta_destino)
        except Exception as e_file:
            conexion.rollback()
            messagebox.showerror("Error", f"No se pudo guardar el comprobante:\n{e_file}", parent=ventana_padre)
            return

        conexion.commit()

        messagebox.showinfo("Éxito", f"La solicitud {id_solicitud} fue marcada como 'Pagado' y se guardó el comprobante.", parent=ventana_padre)
        cargar_solicitudes(tree)

    except Exception as e:
        if conexion: conexion.rollback()
        messagebox.showerror("Error", f"No se pudo actualizar el estado:\n{e}", parent=ventana_padre)
    finally:
        if cursor: cursor.close()
        if conexion: conexion.close()

# Función principal para generar el Excel desde la selección del Treeview
def generar_excel_desde_seleccion(tree, entry_concepto, entry_referencia, entry_factura):
    #id_solicitud = entry_consecutivo.get().strip()
    concepto = entry_concepto.get("1.0", "end").strip()
    referencia_pago = entry_referencia.get().strip()
    global factura_seleccionada

    if not factura_seleccionada:
        messagebox.showerror("Error", "Debes seleccionar un archivo PDF como factura.")
        return

    selected = tree.selection()
    if not selected:
        messagebox.showwarning("Atención", "Seleccione una autorización.")
        return

    id_autorizacion = tree.item(selected[0], "values")[0]

    try:

        conexion = conectar_bd()
        cursor = conexion.cursor()

        # Obtener datos de la autorización
        cursor.execute("""
            SELECT fecha_solicitud, monto, instruccion, id_proveedor, fecha_limite_pago, IVA, subtotal, moneda
            FROM autorizacionescompra 
            WHERE id_autorizacion = %s
        """, (id_autorizacion,))
        autorizacion = cursor.fetchone()
        if not autorizacion:
            messagebox.showerror("Error", "No se encontró la autorización.")
            return

        fecha_solicitud, monto, instruccion, id_proveedor, fechalimite, iva, subtotal, moneda = autorizacion

        # Obtener datos del proveedor
        cursor.execute("""
            SELECT nombre, rfc, email, clave_bancaria, cuenta_bancaria, banco
            FROM proveedores
            WHERE id_proveedor = %s
        """, (id_proveedor,))
        proveedor = cursor.fetchone()
        if not proveedor:
            messagebox.showerror("Error", "No se encontró el proveedor.")
            return

    # Insertar en la tabla SolicitudesPago (sin id_solicitud, MySQL lo genera)
        query = """
            INSERT INTO SolicitudesPago (
                id_autorizacion, id_proveedor, fecha_solicitud, 
                importe, instruccion, referencia_pago, concepto, 
                fecha_recibido_revision, fecha_limite_pago, num_facturas, estado, IVA, subtotal, moneda
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'Pendiente', %s, %s, %s)
        """
        cursor.execute(query, (
        id_autorizacion, id_proveedor, fecha_solicitud,
        monto, instruccion, referencia_pago, concepto, fecha_solicitud, fechalimite,
        factura, iva, subtotal, moneda   # 🔹 aquí guardamos solo el nombre
        ))

        conexion.commit()
        #Obtener id_solicitud
        id_solicitud = cursor.lastrowid

        # Guardar el archivo PDF en la carpeta de facturas con el nombre correcto
        base_onedrive = os.path.join(
            os.environ['USERPROFILE'],
            'OneDrive - ATI',
            'Documentos de AppGestor'
        )
        CARPETA_FACTURAS = os.path.join(base_onedrive, 'Facturas')
        if not os.path.exists(CARPETA_FACTURAS):
            os.makedirs(CARPETA_FACTURAS)

        nombre_archivo = f"Factura_Solicitud de Pago_{id_solicitud}.pdf"
        ruta_destino = os.path.join(CARPETA_FACTURAS, nombre_archivo)

        try:
            shutil.copy(factura_seleccionada, ruta_destino)
            factura = nombre_archivo   # 🔹 SOLO guardamos el nombre en la BD
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo copiar el archivo: {e}")
            return
        
        # Insertar los contratos asociados
        cursor.execute("SELECT id_contrato, importe FROM Autorizacion_Contratos WHERE id_autorizacion = %s", (id_autorizacion,))
        contratos = cursor.fetchall()
        for id_contrato, importe in contratos:
            cursor.execute("""
                INSERT INTO solicitud_contratos (id_solicitud, id_contrato, importe)
                VALUES (%s, %s, %s)
            """, (id_solicitud, id_contrato, importe))
        conexion.commit()

        # Limpiar campos y actualizar UI
        messagebox.showinfo("✅ Éxito", f"Solicitud '{id_solicitud}' guardada en la base de datos.")
        entry_referencia.delete(0, tk.END)
        entry_concepto.delete("1.0", "end")
        entry_factura.config(state="normal")
        entry_factura.delete(0, tk.END)
        entry_factura.config(state="readonly")
        factura_seleccionada = None
        cargar_autorizaciones(tree)

    except mysql.connector.Error as err:
        messagebox.showerror("Error", f"No se pudo registrar la solicitud:\n{err}")
    finally:
        if cursor: cursor.close()
        if conexion: conexion.close()

    # Generar Excel y PDF
    generar_excel(id_solicitud, fecha_solicitud, monto, instruccion,
                  referencia_pago, fechalimite, concepto, factura,
                  *proveedor, usuario_actual["nombre"], iva, subtotal, moneda)

def seleccionar_factura(entry_factura):
    global factura_seleccionada
    if entry_factura.get().strip():
        messagebox.showwarning("Advertencia", "Ya se ha ingresado un número de factura.")
        return

    ruta_origen = filedialog.askopenfilename(
        title="Seleccionar Factura",
        initialdir=os.environ['USERPROFILE'],  # Carpeta personal del usuario
        filetypes=[("Archivos PDF", "*.pdf")]
    )
    if not ruta_origen:
        return

    if ruta_origen:
        factura_seleccionada = ruta_origen
        entry_factura.delete(0, tk.END)
        entry_factura.insert(0, ruta_origen)
        entry_factura.config(state='readonly')

        messagebox.showinfo("Archivo Seleccionado", f"Factura seleccionada:\n{os.path.basename(ruta_origen)}")

def eliminar_factura(entry_factura):
   global factura_seleccionada
   factura_seleccionada = None
   entry_factura.config(state='normal')  # Permitir edición
   entry_factura.delete(0, tk.END)       # Borrar contenido

   seleccionar_factura(entry_factura)

def subir_factura_desde_tree(tree_local, ventana_padre):

    selected = tree_local.selection()
    if not selected:
        messagebox.showwarning("Advertencia", "Debes seleccionar una solicitud.", parent=ventana_padre)
        return

    id_solicitud = tree_local.item(selected[0], "values")[0]

    ruta_origen = filedialog.askopenfilename(
        parent=ventana_padre,
        title="Seleccionar archivo PDF de factura",
        initialdir=os.environ['USERPROFILE'],  # Carpeta personal del usuario
        filetypes=[("Archivos PDF", "*.pdf")]
    )
    if not ruta_origen:
        return

    if not ruta_origen.lower().endswith(".pdf"):
        messagebox.showerror("Error", "Solo se permiten archivos PDF.", parent=ventana_padre)
        return

    try:
        base_onedrive = os.path.join(
            os.environ['USERPROFILE'],
            'OneDrive - ATI',
            'Documentos de AppGestor',
            'Facturas'
        )
        if not os.path.exists(base_onedrive):
            os.makedirs(base_onedrive)

        nombre_archivo = f"Factura_Solicitud de Pago_{id_solicitud}.pdf"
        ruta_destino = os.path.join(base_onedrive, nombre_archivo)

        # Eliminar si ya existía
        if os.path.exists(ruta_destino):
            os.remove(ruta_destino)

        # Copiar nuevo archivo
        shutil.copy(ruta_origen, ruta_destino)

        # Actualizar la base de datos
        conexion = conectar_bd()
        cursor = conexion.cursor()
        cursor.execute("""
            UPDATE SolicitudesPago
            SET num_facturas = %s
            WHERE id_solicitud = %s
        """, (ruta_destino, id_solicitud))
        conexion.commit()
        cursor.close()
        conexion.close()

        messagebox.showinfo("✅ Éxito", f"Factura actualizada para la solicitud {id_solicitud}.", parent=ventana_padre)

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo subir la factura:\n{e}", parent=ventana_padre)

# Función que llena la plantilla Excel con los datos
def generar_excel(id_solicitud, fecha_solicitud, monto, instruccion,
                  referencia_pago, fechalimite, concepto, factura, nombre, rfc, email, clave_bancaria, cuenta_bancaria, banco, nombre_usuario, iva, subtotal, moneda):
  
    try:
        # Obtener los nombres de contratos asociados a la solicitud
        conexion = conectar_bd()
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT c.contrato 
            FROM solicitud_contratos sc
            JOIN contratos c ON sc.id_contrato = c.id_contrato
            WHERE sc.id_solicitud = %s
        """, (id_solicitud,))
        nombres_contratos = [fila[0] for fila in cursor.fetchall()]
        texto_contratos = ", ".join(nombres_contratos)

        # Plantilla de solicitud de pago
        plantilla_path = ruta_relativa("Plantillas/Solicitud_Pago.xlsx")
        wb = load_workbook(plantilla_path)
        sheet = wb.active

        # Función para escribir en celdas combinadas
        def escribir(fila, columna, valor, combinar=None):
            celda = sheet.cell(row=fila, column=columna)

            # Verificar si está en un rango combinado
            for r in sheet.merged_cells.ranges:
                if celda.coordinate in r:
                    sheet.unmerge_cells(str(r))
                    break

            celda.value = valor
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text= True)

            # Si se desea recombinar
            if combinar:
                sheet.merge_cells(combinar)

        # Llenar celdas con datos generales
        escribir(6, 10, id_solicitud, combinar="J6:L6")          # J6 - Consecutivo
        escribir(9, 3, fecha_solicitud, combinar="C9:E9")        # C9 - Fecha
        escribir(9, 8, subtotal, combinar="H9:I9")               # Subtotal
        escribir(9, 10, iva, combinar="J9:K9")                   # IVA
        escribir(9, 12, f"{monto} {moneda}", combinar="L9:L9")  # Total con moneda
        escribir(34, 8, texto_contratos, combinar="H34:L34")     # H34 - Proyecto

        escribir(12, 3, nombre, combinar="C12:L12")              # C12 - Nombre proveedor
        escribir(15, 7, rfc, combinar="G15:L15")                 # G15 - RFC
        escribir(18, 8, email, combinar="H18:L18")               # G18 - Email
        escribir(18, 3, clave_bancaria, combinar="C18:F18")      # C18 - Clave bancaria
        escribir(22, 3, cuenta_bancaria, combinar="C22:E22")     # C22 - Cuenta bancaria
        escribir(22, 7, banco, combinar="G22:H22")               # G22 - Banco
        escribir(29, 8, fechalimite, combinar="H29:L29")         # H29 - Limite de Pago
        escribir(15, 3, instruccion, combinar="C15:E15")         # C15 - Instrucción
        escribir(22, 10, referencia_pago, combinar="J22:L22")    # J22 - Referencia de pago
        escribir(25, 3, concepto, combinar="C25:L25")            # C25 - Concepto
        escribir(38, 3, nombre_usuario, combinar="C38:F38")      # C37 - Solicitante de Pago
        
        # Ver si la factura es un archivo PDF o texto normal
        if factura.lower().endswith(".pdf"):
            base_onedrive = os.path.join(
                os.environ['USERPROFILE'],
                'OneDrive - ATI',
                'Documentos de AppGestor',
                'Facturas'
            )
            RUTA_FACTURAS = os.path.join(base_onedrive, factura)  # 🔹 Usamos el nombre y construimos ruta
        else:
            RUTA_FACTURAS = factura

        if os.path.exists(RUTA_FACTURAS):
            celda_factura = sheet.cell(row=34, column=3)
            celda_factura.value = factura  # 🔹 Solo el nombre
            celda_factura.hyperlink = RUTA_FACTURAS
            celda_factura.style = "Hyperlink"
            celda_factura.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Definir estilo de borde
            borde_completo = Border(
                top=Side(style="thin"),
                bottom=Side(style="thin"),
                left=Side(style="thin"),
                right=Side(style="thin")
            )
            sheet.merge_cells("C34:F34")
            celda_combinada = sheet["C34"]
            celda_combinada.border = borde_completo
        else:
            escribir(34, 3, factura, combinar="C34:F34")


        ruta_firma = ruta_relativa(usuario_actual["firma"])

        # Insertar imagen
        firma_img = ExcelImage(ruta_firma)
        firma_img.width =160 #ajustar tamaño
        firma_img.height = 50
        sheet.add_image(firma_img, "D37")

        # Crear la carpeta si no existe
        if not os.path.exists(CARPETA_SOLICITUDES):
            os.makedirs(CARPETA_SOLICITUDES)

        # Guardar Excel en carpeta
        nombre_excel = f"Solicitud de Pago_{id_solicitud}.xlsx"
        output_path = os.path.join(CARPETA_SOLICITUDES, nombre_excel)
        wb.save(output_path)

        # Convertir a PDF
        nombre_pdf = f"Solicitud de Pago_{id_solicitud}.pdf"
        ruta_pdf = os.path.join(CARPETA_SOLICITUDES, nombre_pdf)
        convertir_excel_a_pdf(output_path, ruta_pdf)

        # 🔹 Guardar en BD solo el nombre del PDF (no la ruta)
        cursor.execute(
            "UPDATE solicitudespago SET archivo = %s WHERE id_solicitud = %s",
            (nombre_pdf, id_solicitud)
        )
        conexion.commit()

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo generar el Excel:\n{e}")

def mSolicitud_excel(id_autorizacion):
    conexion = conectar_bd()
    cursor = conexion.cursor()

    # Buscar solicitud relacionada
    cursor.execute("SELECT * FROM SolicitudesPago WHERE id_autorizacion = %s", (id_autorizacion,))
    solicitud = cursor.fetchone()
    if not solicitud:
        messagebox.showinfo("Sin solicitud", "No hay una solicitud de pago relacionada a esta autorización.")
        return

    (
        id_solicitud, _, id_proveedor, fecha_solicitud, monto, instruccion,
        referencia_pago, concepto, _, fecha_limite, factura, _, estado, iva, subtotal
    ) = solicitud

    # Buscar datos actuales de autorización y proveedor
    cursor.execute("""
        SELECT a.fecha_solicitud, a.monto, a.instruccion, a.fecha_limite_pago,
               p.nombre, p.rfc, p.email, p.clave_bancaria, p.cuenta_bancaria, p.banco
        FROM autorizacionescompra a
        JOIN proveedores p ON a.id_proveedor = p.id_proveedor
        WHERE a.id_autorizacion = %s
    """, (id_autorizacion,))
    resultado = cursor.fetchone()
    if not resultado:
        messagebox.showerror("Error", "No se encontró la autorización o proveedor.")
        return

    nueva_fecha, nuevo_monto, nueva_instr, nueva_fecha_lim, nombre, rfc, email, clave, cuenta, banco = resultado

    # Comparar campos clave para ver si ha habido cambios
    cambios = (
        str(fecha_solicitud) != str(nueva_fecha) or
        float(monto) != float(nuevo_monto) or
        instruccion != nueva_instr or
        str(fecha_limite) != str(nueva_fecha_lim)
    )

    if not cambios:
        print("No se detectaron cambios en la solicitud de pago. No se modificará el Excel.")
        return

    # Generar nuevo Excel sobrescribiendo
    generar_excel(
        id_solicitud, nueva_fecha, nuevo_monto, nueva_instr, referencia_pago,
        nueva_fecha_lim, concepto, factura, nombre, rfc, email, clave, cuenta,
        banco, usuario_actual["nombre"], iva, subtotal
    )

    print(f"Solicitud de pago {id_solicitud} actualizada.")

    cursor.close()
    conexion.close()

#Ventana de solicitudes pagadas
def solicitudes_pagadas(tree, search_text="", fecha_inicio=None, fecha_fin=None, proveedor=None):
    for row in tree.get_children():
        tree.delete(row)

    conexion = None
    cursor = None
    try:
        conexion = conectar_bd()
        if conexion is None:
            return
        cursor = conexion.cursor()

        query = """
            SELECT s.id_solicitud, s.fecha_solicitud, s.fecha_pago, s.importe,
                   p.nombre, s.estado, s.comprobante
            FROM solicitudespago s
            INNER JOIN proveedores p ON s.id_proveedor = p.id_proveedor
            WHERE s.estado = 'Pagado'
        """
        params = []

        if search_text:
            query += " AND (s.id_solicitud LIKE %s OR p.nombre LIKE %s)"
            params.extend([f"%{search_text}%", f"%{search_text}%"])

        if fecha_inicio and fecha_fin:
            query += " AND s.fecha_pago BETWEEN %s AND %s"
            params.extend([fecha_inicio, fecha_fin])

        if proveedor and proveedor != "Todos":
            id_proveedor = proveedor.split(" - ")[0]
            query += " AND p.id_proveedor = %s"
            params.append(id_proveedor)

        cursor.execute(query, params)
        for row in cursor.fetchall():
            tree.insert("", "end", values=row)

    except mysql.connector.Error as e:
        messagebox.showerror("Error", f"No se pudieron cargar las solicitudes: {e}")
    finally:
        if cursor: cursor.close()
        if conexion: conexion.close()

def ventana_solicitudes_pagadas():
    ventana = tk.Toplevel()
    ventana.title("Solicitudes Pagadas")
    centrar_ventana(ventana, 1200, 600)
    ventana.configure(bg="white")

    # Canvas para degradado
    canvas = tk.Canvas(ventana)
    canvas.place(relx=0, rely=0, relwidth=1, relheight=1)
    def actualizar_degradado(event):
        ancho = canvas.winfo_width()
        alto = canvas.winfo_height()
        crear_degradado_vertical(canvas, ancho, alto, "#8B0000", "#FFFFFF")
    canvas.bind("<Configure>", actualizar_degradado)
    ventana.after(100, lambda: actualizar_degradado(None))

    # Frame de filtros
    frame_filtros = ttk.Frame(canvas, padding=10)
    frame_filtros.pack(fill="x")

    # --- Filtros ---
    tk.Label(frame_filtros, text="Proveedor:", font=("Arial", 10, "bold")).pack(side="left", padx=5)
    combo_proveedor = ttk.Combobox(frame_filtros, width=40, state="readonly")
    combo_proveedor.pack(side="left", padx=5)

    # Cargar proveedores
    conexion = conectar_bd()
    cursor = conexion.cursor()
    cursor.execute("SELECT DISTINCT nombre FROM proveedores ORDER BY nombre ASC")
    proveedores = [row[0] for row in cursor.fetchall()]
    conexion.close()
    combo_proveedor['values'] = ["Todos los proveedores"] + proveedores
    combo_proveedor.current(0)

    # Fechas
    tk.Label(frame_filtros, text="Fecha inicio:").pack(side="left", padx=5)
    entry_fecha_inicio = DateEntry(frame_filtros, date_pattern='yyyy-mm-dd', width=12,
                                   background='darkblue', foreground='white', borderwidth=2)
    entry_fecha_inicio.pack(side="left", padx=5)

    tk.Label(frame_filtros, text="Fecha fin:").pack(side="left", padx=5)
    entry_fecha_fin = DateEntry(frame_filtros, date_pattern='yyyy-mm-dd', width=12,
                                background='darkblue', foreground='white', borderwidth=2)
    entry_fecha_fin.pack(side="left", padx=5)

    # Barra de búsqueda
    tk.Label(frame_filtros, text="Buscar:").pack(side="left", padx=5)
    entry_busqueda_var = tk.StringVar()
    entry_busqueda = ttk.Entry(frame_filtros, textvariable=entry_busqueda_var)
    entry_busqueda.pack(side="left", padx=5)

    # --- TREEVIEW ---
    tree = ttk.Treeview(ventana, columns=("id", "fecha_solicitud", "fecha_pago", "importe", "proveedor", "estado", "comprobante"), show="headings", height=15)
    tree.heading("id", text="No. Solicitud", anchor="center")
    tree.heading("fecha_solicitud", text="Fecha Solicitud", anchor="center")
    tree.heading("fecha_pago", text="Fecha Pago", anchor="center")
    tree.heading("importe", text="Importe", anchor="center")
    tree.heading("proveedor", text="Proveedor", anchor="center")
    tree.heading("estado", text="Estado", anchor="center")
    tree.heading("comprobante", text="Comprobante", anchor="center")

    tree.column("id", width=90, anchor="center")
    tree.column("fecha_solicitud", width=120, anchor="center")
    tree.column("fecha_pago", width=120, anchor="center")
    tree.column("importe", width=100, anchor="center")
    tree.column("proveedor", width=200, anchor="w")
    tree.column("estado", width=100, anchor="center")
    tree.column("comprobante", width=300, anchor="w")

    tree.place(relx=0.03, rely=0.25, relwidth=0.9, relheight=0.6)

    # --- Funciones ---
    def cargar_solicitudes():
        """Carga solicitudes según filtros de proveedor y fechas"""
        for item in tree.get_children():
            tree.delete(item)

        proveedor = combo_proveedor.get()
        fecha_inicio = entry_fecha_inicio.get()
        fecha_fin = entry_fecha_fin.get()

        conexion = conectar_bd()
        cursor = conexion.cursor()
        query = """
            SELECT sp.id_solicitud, sp.fecha_solicitud, sp.fecha_pago, sp.importe,
                   p.nombre, sp.estado, sp.comprobante
            FROM SolicitudesPago sp
            JOIN proveedores p ON sp.id_proveedor = p.id_proveedor
            WHERE sp.estado = 'Pagado' AND DATE(sp.fecha_pago) BETWEEN %s AND %s
        """
        params = [fecha_inicio, fecha_fin]
        if proveedor != "Todos los proveedores":
            query += " AND p.nombre = %s"
            params.append(proveedor)

        cursor.execute(query, tuple(params))
        for row in cursor.fetchall():
            tree.insert("", tk.END, values=row)
        conexion.close()

    def buscar(*args):
        termino = entry_busqueda_var.get().strip().lower()
        for item in tree.get_children():
            tree.delete(item)

        conexion = conectar_bd()
        cursor = conexion.cursor()

        if termino:
            consulta = """
                SELECT sp.id_solicitud, sp.fecha_solicitud, sp.fecha_pago, sp.importe,
                       p.nombre, sp.estado, sp.comprobante
                FROM SolicitudesPago sp
                JOIN proveedores p ON sp.id_proveedor = p.id_proveedor
                WHERE sp.estado = 'Pagado' AND (
                    CAST(sp.id_solicitud AS CHAR) LIKE %s OR
                    CAST(sp.fecha_pago AS CHAR) LIKE %s OR
                    CAST(sp.importe AS CHAR) LIKE %s OR
                    LOWER(p.nombre) LIKE %s
                )
            """
            like_termino = f"%{termino}%"
            params = [like_termino]*4
        else:
            cargar_solicitudes()
            return

        cursor.execute(consulta, tuple(params))
        for row in cursor.fetchall():
            tree.insert("", tk.END, values=row)
        conexion.close()

    # Activar búsqueda en vivo
    entry_busqueda_var.trace_add("write", buscar)

    # Botón abrir comprobante
    def abrir_comprobante_sel():
        item = tree.focus()
        if not item:
            messagebox.showwarning("Atención", "Selecciona una solicitud.", parent=ventana)
            return

        nombre_archivo = tree.item(item, "values")[6]  # Contiene solo el nombre (ej: Comprobante_123.pdf)

        if not nombre_archivo:
            messagebox.showerror("Error", "No se encontró el nombre del comprobante.", parent=ventana)
            return

        # Construir la ruta completa en tiempo de ejecución
        ruta_completa = os.path.join(CARPETA_COMPROBANTES, nombre_archivo)

        if os.path.exists(ruta_completa):
            os.startfile(ruta_completa)  # Solo para Windows
        else:
            messagebox.showerror("Error", f"No se encontró el comprobante en:\n{ruta_completa}", parent=ventana)

    # Botón en la interfaz
    ttk.Button(frame_filtros, text="Abrir Comprobante", command=abrir_comprobante_sel).pack(side="left", padx=10)
    ttk.Button(frame_filtros, text="Aplicar Filtros", command=cargar_solicitudes).pack(side="left", padx=10)
    tk.Button(canvas, text="Salir", command= ventana.destroy, bg="red", fg="white", font=("Arial", 10, "bold")).place(relx=0.1, rely=0.91, relwidth=0.095, relheight=0.05)
    

    # Carga inicial
    cargar_solicitudes()

# Interfaz gráfica
def hex_a_rgb(hex_color):
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

def crear_degradado_vertical(canvas, ancho, alto, color_inicio, color_fin):
    canvas.delete("degradado")  # Limpiar el canvas antes de dibujar un nuevo degradado

    rgb_inicio = hex_a_rgb(color_inicio)
    rgb_fin = hex_a_rgb(color_fin)

    pasos = alto // 2
    for i in range(pasos):
        y = alto - i - 1
        r = int(rgb_inicio[0] + (rgb_fin[0] - rgb_inicio[0]) * i / pasos)
        g = int(rgb_inicio[1] + (rgb_fin[1] - rgb_inicio[1]) * i / pasos)
        b = int(rgb_inicio[2] + (rgb_fin[2] - rgb_inicio[2]) * i / pasos)
        color = f"#{r:02x}{g:02x}{b:02x}"
        canvas.create_line(0, y, ancho, y, fill=color, tags="degradado")

    canvas.create_rectangle(0, 0, ancho, alto // 2, fill=color_fin, outline="", tags="degradado")


def gestionar_solicitudes(rol,volver_menu_callback):
    ventana = tk.Tk()
    ventana.title("Solicitudes de Pago")
    centrar_ventana(ventana, 1020, 600)

    canvas = tk.Canvas(ventana)
    canvas.pack(fill="both", expand=True)


    def actualizar_degradado(event):
        # Obtener las dimensiones del canvas
        ancho = canvas.winfo_width()
        alto = canvas.winfo_height()
        crear_degradado_vertical(canvas, ancho, alto, "#8B0000", "#FFFFFF")

    # Actualizar el fondo degradado al cambiar el tamaño de la ventana
    canvas.bind("<Configure>", actualizar_degradado)

    # Inicializar el degradado en el tamaño actual de la ventana
    ventana.after(100, lambda: actualizar_degradado(None))

    RUTA_LOGO = ruta_relativa("Plantillas/LogoATM.png")
    RUTA_LOGO2 = ruta_relativa("Plantillas/ISO-9001.jpeg")
    RUTA_LOGO3 = ruta_relativa("Plantillas/ISO-14001.jpeg")
    RUTA_LOGO4 = ruta_relativa("Plantillas/ISO-45001.jpeg")
    try:
        # LOGO ATM
        imagen = Image.open(RUTA_LOGO)
        imagen = imagen.resize((120, 160), Image.Resampling.LANCZOS)
        logo_img = ImageTk.PhotoImage(imagen)
        label_logo = tk.Label(canvas, image=logo_img, borderwidth=0)
        label_logo.image = logo_img
        label_logo.place(relx=0.05, rely=0.015)

        # ISO 9001
        iso1 = Image.open(RUTA_LOGO2).resize((60, 60), Image.Resampling.LANCZOS)
        iso1_img = ImageTk.PhotoImage(iso1)
        label_iso1 = tk.Label(canvas, image=iso1_img, borderwidth=0, bg="#ffffff")
        label_iso1.image = iso1_img
        label_iso1.place(relx=0.30, rely=0.020, anchor="ne")  # Esquina inferior derecha

        # ISO 14001
        iso2 = Image.open(RUTA_LOGO3).resize((60, 60), Image.Resampling.LANCZOS)
        iso2_img = ImageTk.PhotoImage(iso2)
        label_iso2 = tk.Label(canvas, image=iso2_img, borderwidth=0, bg="#ffffff")
        label_iso2.image = iso2_img
        label_iso2.place(relx=0.35, rely=0.13, anchor="ne")  # Al lado izquierdo del ISO 9001

        # ISO 45001
        iso3 = Image.open(RUTA_LOGO4).resize((60, 60), Image.Resampling.LANCZOS)
        iso3_img = ImageTk.PhotoImage(iso3)
        label_iso3 = tk.Label(canvas, image=iso3_img, borderwidth=0, bg="#ffffff")
        label_iso3.image = iso3_img
        label_iso3.place(relx=0.25, rely=0.13, anchor="ne")  # Al lado izquierdo del ISO 14001

    except Exception as e:
        print(f"⚠️ No se pudo cargar el logotipo: {e}")
        label_logo = tk.Label(canvas, text="LOGO", font=("Arial", 20, "bold"))


    # Buscar
    tk.Label(ventana, text="Buscar:", font=("Arial", 10, "bold"), bg="white").place(relx=0.05, rely=0.30)
    entry_busqueda = tk.Entry(ventana, width=50)
    entry_busqueda.place(relx=0.13, rely=0.30)

    def buscar(*args):  # Acepta *args por el trace
        termino = entry_busqueda.get().lower()
        for item in tree.get_children():
            tree.delete(item)

        conexion = conectar_bd()
        cursor = conexion.cursor()
        if termino:
            consulta = """
                SELECT id_autorizacion, fecha_solicitud, monto FROM autorizacionescompra
                WHERE LOWER(id_autorizacion) LIKE %s
                OR LOWER(fecha_solicitud) LIKE %s
                OR LOWER(monto) LIKE %s       
            """
            like_termino = f"%{termino}%"
            cursor.execute(consulta, (like_termino, like_termino, like_termino))

        else : 
            consulta= ("""
                SELECT id_autorizacion, fecha_solicitud, monto
                FROM autorizacionescompra
                WHERE id_autorizacion NOT IN (
                    SELECT id_autorizacion FROM SolicitudesPago
                )""")
            cursor.execute(consulta)
            
        resultados = cursor.fetchall()
        conexion.close()

        for row in resultados:
            tree.insert("", tk.END, values=row)

    entry_busqueda_var = tk.StringVar()
    entry_busqueda.config(textvariable=entry_busqueda_var)
    entry_busqueda_var.trace("w", buscar)  # Búsqueda automática al escribir

    tk.Label(ventana, text="Ingrese la informacion de la solicitud:", font=("Arial", 12, "bold"), bg="white").place(relx=0.60, rely=0.02)

    # Concepto
    tk.Label(ventana, text="Concepto:", font=("Arial", 10, "bold"), bg="white").place(relx=0.60, rely=0.16)
    entry_concepto = tk.Text(ventana, wrap="word", font=("Arial", 10))
    entry_concepto.place(relx=0.70, rely=0.14, relwidth=0.23, relheight=0.06)

    # Referencia de Pago
    tk.Label(ventana, text="Referencia de Pago:", font=("Arial", 10, "bold"), bg="white").place(relx=0.60, rely=0.22)
    entry_referencia = tk.Entry(ventana, width=30)
    entry_referencia.place(relx=0.75, rely=0.22)

    #Numero de Factura
    tk.Label(ventana, text="Numero de Factura:", font=("Arial", 10, "bold"), bg="white").place(relx=0.60, rely=0.28)
    entry_factura = tk.Entry(ventana,width=15)
    entry_factura.place(relx=0.75, rely=0.28)
    tk.Button(ventana, text="Cargar Factura/Cotizacion", command=lambda: seleccionar_factura(entry_factura)).place(relx=0.85, rely=0.28)
    tk.Button(ventana, text= "Cambiar Factura", command=lambda: eliminar_factura(entry_factura)).place(relx=0.85, rely=0.33)


    #Aplicacion del estilo a la tabla
    style = ttk.Style()
    style.theme_use("alt")
    style.configure("Treeview.Heading", font=("Arial", 10, "bold"), foreground="white", background="#990000")
    style.map("Treeview.Heading", background=[("!active", "#990000"), ("active", "#990000"), ("pressed", "#990000")],
              foreground=[("!active", "white"), ("active", "white"), ("pressed", "white")])

    # Treeview (tabla)
    tree = ttk.Treeview(ventana, columns=("ID", "Fecha", "Monto", "Limite de Pago"), show="headings")
    for col in ("ID", "Fecha", "Monto", "Limite de Pago"):
        tree.heading(col, text=col)
        tree.column(col, anchor="center")

    # Scrollbar
    scrollbar = ttk.Scrollbar(ventana, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)

    tree.place(relx=0.05, rely=0.38, relwidth=0.88, relheight=0.50)
    scrollbar.place(relx=0.93, rely=0.38, relheight=0.50)

    # Ventana de solicitudes guardadas
    def Solicitudes(parent=None):
        ventana_solicitudes = tk.Toplevel(parent) if parent else tk.Toplevel()
        ventana_solicitudes.title("Solicitudes Guardadas")
        centrar_ventana(ventana_solicitudes, 1200, 650)  # un poco más grande

        #Aplicacion del estilo a la tabla
        style = ttk.Style()
        style.theme_use("alt")
        style.configure("Treeview.Heading", font=("Arial", 10, "bold"), foreground="white", background="#990000")
        style.map("Treeview.Heading", background=[("!active", "#990000"), ("active", "#990000"), ("pressed", "#990000")],
                foreground=[("!active", "white"), ("active", "white"), ("pressed", "white")])

        frame_tabla = tk.Frame(ventana_solicitudes)
        frame_tabla.place(relx=0.05, rely=0.05, relwidth=0.9, relheight=0.75)

        # Agrego columna "Archivo" para guardar el nombre del PDF
        tree_local = ttk.Treeview(
            frame_tabla,
            columns=("ID", "Fecha", "Importe", "Limite de Pago", "Estado", "Archivo"),
            show="headings"
        )

        # Configuración de columnas
        tree_local.heading("ID", text="No. Solicitud")
        tree_local.column("ID", width=90, anchor="center")

        tree_local.heading("Fecha", text="Fecha")
        tree_local.column("Fecha", width=100, anchor="center")

        tree_local.heading("Importe", text="Importe")
        tree_local.column("Importe", width=100, anchor="center")

        tree_local.heading("Limite de Pago", text="Límite de Pago")
        tree_local.column("Limite de Pago", width=120, anchor="center")

        tree_local.heading("Estado", text="Estado")
        tree_local.column("Estado", width=100, anchor="center")

        tree_local.heading("Archivo", text="Solicitud PDF")
        tree_local.column("Archivo", width=350, anchor="w")  # ⬅️ izquierda porque es un nombre largo

        tree_local.place(relx=0, rely=0, relwidth=0.97, relheight=1)

        scrollbar_local = ttk.Scrollbar(frame_tabla, orient="vertical", command=tree_local.yview)
        tree_local.configure(yscrollcommand=scrollbar_local.set)
        scrollbar_local.place(relx=0.97, rely=0, relwidth=0.03, relheight=1)

        # Buscar por ID
        tk.Label(ventana_solicitudes, text="Buscar ID:", font=("Arial", 10, "bold")).place(relx=0.1, rely=0.85)
        entry_busqueda = ttk.Entry(ventana_solicitudes, width=20)
        entry_busqueda.place(relx=0.2, rely=0.85)

        def abrir_solicitud_sel():
            item = tree_local.focus()
            if not item:
                messagebox.showwarning("Atención", "Selecciona una solicitud.", parent=ventana_solicitudes)
                return

            valores = tree_local.item(item, "values")
            if not valores or len(valores) < 6:
                messagebox.showerror("Error", "No se encontró el nombre de la solicitud.", parent=ventana_solicitudes)
                return

            nombre_archivo = valores[5]  # columna "Archivo"
            if not nombre_archivo:
                messagebox.showerror("Error", "No se encontró el nombre del archivo.", parent=ventana_solicitudes)
                return

            # Construir la ruta completa en tiempo de ejecución
            ruta_completa = os.path.join(CARPETA_SOLICITUDES, nombre_archivo)

            if os.path.exists(ruta_completa):
                try:
                    os.startfile(ruta_completa)  # Windows
                except Exception as e:
                    messagebox.showerror("Error", f"No se pudo abrir la solicitud:\n{e}", parent=ventana_solicitudes)
            else:
                messagebox.showerror("Error", f"No se encontró la solicitud en:\n{ruta_completa}", parent=ventana_solicitudes)
        # 📌 También abrir con doble clic en la fila
        tree_local.bind("<Double-1>", lambda e: abrir_solicitud_sel())

        def buscar_solicitud():
            id_buscar = entry_busqueda.get().strip()
            for item in tree_local.get_children():
                tree_local.selection_remove(item)
                valores = tree_local.item(item, "values")
                if valores and str(valores[0]) == id_buscar:
                    tree_local.see(item)
                    tree_local.selection_add(item)
                    break
            else:
                ventana_solicitudes.lift()
                ventana_solicitudes.focus_force()
                messagebox.showinfo("No encontrado", f"No se encontró la solicitud con ID {id_buscar}")

        # Botones
        tk.Button(ventana_solicitudes, text="Buscar", command=buscar_solicitud,
                font=("Arial", 10, "bold"), bg="#004080", fg="white").place(relx=0.35, rely=0.845)

        tk.Button(ventana_solicitudes, text="Salir", command=ventana_solicitudes.destroy,
                bg="red", fg="white", font=("Arial", 10, "bold")).place(relx=0.1, rely=0.92, relwidth=0.08, relheight=0.04)

        tk.Button(ventana_solicitudes, text="Marcar como Pagado", font=("Arial", 10, "bold"),
                command=lambda: marcar_como_pagado(tree_local, usuario_actual, ventana_solicitudes),
                bg="#006400", fg="white").place(relx=0.6, rely=0.92, relwidth=0.2, relheight=0.06)

        tk.Button(
            ventana_solicitudes,
            text="📤Subir/Actualizar Factura",
            bg="#004080", fg="white",
            font=("Arial", 10, "bold"),
            command=lambda: subir_factura_desde_tree(tree_local, ventana_solicitudes)
        ).place(relx=0.6, rely=0.84, relwidth=0.2, relheight=0.06)

        tk.Button(ventana_solicitudes, text="Control de Soliciudes", command=lambda: Control_Solicitudes(usuario_actual, ventana_solicitudes), font=("Arial", 10, "bold")
                ).place(relx=0.83, rely=0.92, relwidth=0.15, relheight=0.06)

        tk.Button(ventana_solicitudes, text="Solicitudes Pagadas", bg="green", fg="white",
                command=lambda: ventana_solicitudes_pagadas(), font=("Arial", 10, "bold")
                ).place(relx=0.83, rely=0.84, relwidth=0.15, relheight=0.06)

        # Cargar solicitudes en el Treeview
        cargar_solicitudes(tree_local)

    # Botones
    tk.Button(ventana, text="Guardar y Generar Docs",
              command=lambda: generar_excel_desde_seleccion(tree, entry_concepto, entry_referencia, entry_factura), font=("Arial", 10, "bold")
             ).place(relx=0.35, rely=0.91, relwidth=0.18, relheight=0.05)

    tk.Button(ventana, text="Solicitudes Guardadas", command=Solicitudes, font=("Arial", 10, "bold")).place(relx=0.75, rely=0.91, relwidth=0.15, relheight=0.05)
    tk.Button(ventana, text="Salir", command= lambda: salir(volver_menu_callback, ventana), bg="red", fg="white", font=("Arial", 10, "bold")).place(relx=0.1, rely=0.91, relwidth=0.095, relheight=0.05)
    tk.Button(ventana, text="Reporte de Costos", command=lambda: costos_contrato(), font=("Arial", 10, "bold")
              ).place(relx=0.55, rely=0.91, relwidth=0.15, relheight=0.05)

    cargar_autorizaciones(tree)
    ventana.mainloop()