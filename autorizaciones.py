import shutil
import tkinter as tk
from tkinter import ttk, messagebox
from tkinter import filedialog
from tkcalendar import DateEntry
from datetime import datetime
import mysql.connector
from proveedores import cargar_proveedores, ventana_agregar_proveedor
from solicitudes import mSolicitud_excel
from utils import ruta_relativa, centrar_ventana, convertir_excel_a_pdf, salir
from login import usuario_actual
from database import conectar_bd
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

#Funcion para agregar las autorizaciones
def agregar_autorizacion(entry_consecutivo, combo_tipo, combo_solicitante, entry_puesto, entry_area, entry_fecha_solicitud,
                         entry_fecha_requerida, entry_monto, combo_proveedor, combo_instruccion,
                         entry_flimite, tree, listbox_contratos, entry_IVA, entry_subtotal, combo_moneda, entry_general):
    
    consecutivo = entry_consecutivo.get()
    tipo = combo_tipo.get()
    solicitante = combo_solicitante.get()
    puesto = entry_puesto.get()
    area = entry_area.get()
    # Convertir a objeto fecha y luego a string en formato correcto
    fecha_solicitud = datetime.strptime(entry_fecha_solicitud.get(), "%Y/%m/%d").strftime("%Y-%m-%d")
    fecha_requerida = datetime.strptime(entry_fecha_requerida.get(), "%Y/%m/%d").strftime("%Y-%m-%d")
    monto = entry_monto.get()
    id_proveedor = combo_proveedor.get().split(" - ")[0]
    instruccion = combo_instruccion.get()
    limite = entry_flimite.get()
    seleccionados = listbox_contratos.curselection()
    iva = entry_IVA.get()
    subtotal = entry_subtotal.get()
    moneda = combo_moneda.get()
    general = entry_general.get("1.0", "end").strip()

    def seleccionar_cotizacion(consecutivo):
        # Selección del archivo original (cotización)
        ruta_origen = filedialog.askopenfilename(
        title="Seleccionar Cotizacion",
        initialdir=os.environ['USERPROFILE'],  # Carpeta personal del usuario
        filetypes=[("Archivos PDF", "*.pdf")]
        ) 

        if ruta_origen:
                # Ruta base dentro de OneDrive
                base_onedrive = os.path.join(
                    os.environ['USERPROFILE'],
                    'OneDrive - ATI',
                    'Documentos de AppGestor'
                )

                # Carpeta destino para cotizaciones compartidas
                CARPETA_COTIZACIONES = os.path.join(base_onedrive, 'Cotizaciones')

                # Crear la carpeta si no existe
                os.makedirs(CARPETA_COTIZACIONES, exist_ok=True)

                # Nombre y ruta destino
                nombre_destino = f"Cotizacion_Autorizacion de Compra_{consecutivo}.pdf"
                ruta_destino = os.path.join(CARPETA_COTIZACIONES, nombre_destino)

                # Copiar archivo
                shutil.copy(ruta_origen, ruta_destino)

                # Devolver ruta con slashes compatibles
                return ruta_destino.replace("\\", "/")
        return None

    if not (tipo and solicitante and puesto and area and fecha_solicitud and fecha_requerida and monto and id_proveedor and instruccion):
        messagebox.showwarning("Campos vacíos", "Por favor, llena todos los campos.")
        return

    if not articulos_lista:
        messagebox.showwarning("Campos vacíos", "Debe agregar al menos un artículo antes de registrar la autorización.")
        return
    
    if not seleccionados:
        messagebox.showwarning("Campos vacíos", "Selecciona al menos un contrato.")
        return    
    
    if moneda not in ("MXN", "USD"):
        messagebox.showerror("Error", "Selecciona una moneda válida.")
        return
    
    # ✅ Insertar los contratos relacionados
    try:
        monto_float = float(monto)
    except ValueError:
        messagebox.showerror("Error", "El monto total ingresado no es válido.")
        return

    # Obtener contratos seleccionados
    contratos_ids_nombres = [(listbox_contratos.get(i).split(" - ")[0], listbox_contratos.get(i)) for i in seleccionados]

    # Obtener importes desde porcentajes
    importes_contratos = asignar_importes_a_contratos(tree.winfo_toplevel(), contratos_ids_nombres, monto_float)

    if importes_contratos is None:
        messagebox.showinfo("Cancelado", "Registro de autorización cancelado por el usuario.")
        return  # Detener el guardado

    conexion = None
    cursor = None

    try:
        conexion = conectar_bd()
        if conexion is None:
            print("❌No se pudo establecer la conexion")
            return
        cursor = conexion.cursor()

        cursor.execute("SELECT COUNT(*) FROM AutorizacionesCompra WHERE id_autorizacion = %s", (consecutivo,))
        if cursor.fetchone()[0] > 0:
            messagebox.showerror("Error", f"El ID de autorización '{consecutivo}' ya existe. Elija otro.")
            return

        query = """
        INSERT INTO AutorizacionesCompra 
        (id_autorizacion, tipo_solicitud, solicitante, puesto, area, fecha_solicitud, 
         fecha_requerida, monto, id_proveedor, instruccion, fecha_limite_pago, estado, IVA, subtotal, moneda, generales)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'Pendiente', %s, %s, %s, %s)
        """
        valores = (consecutivo, tipo, solicitante, puesto, area, fecha_solicitud, fecha_requerida,
                  monto, id_proveedor, instruccion, limite, iva, subtotal, moneda, general)
        cursor.execute(query, valores)

        ruta_cotizacion = seleccionar_cotizacion(consecutivo)

        # Inserta/actualiza en la tabla AutorizacionesCompra
        cursor.execute("""
        UPDATE AutorizacionesCompra
        SET ruta_cotizacion = %s
        WHERE id_autorizacion = %s
    """, (ruta_cotizacion, consecutivo))

        conexion.commit()

        # Insertar en tabla con importes asignados
        for id_contrato, importe in importes_contratos.items():
            cursor.execute(
                "INSERT INTO Autorizacion_Contratos (id_autorizacion, id_contrato, importe) VALUES (%s, %s, %s)",
                (consecutivo, id_contrato, importe)
            )

        # ✅ Insertar los artículos relacionados
        query_articulo = """
            INSERT INTO ArticulosAutorizacion (id_autorizacion, cantidad, unidad, articulo, observaciones)
            VALUES (%s, %s, %s, %s, %s)
        """
        for articulo in articulos_lista:
            cursor.execute(query_articulo, (consecutivo, *articulo))

        conexion.commit()
        messagebox.showinfo("✅Éxito", "Autorización y artículos registrados correctamente.")

    except mysql.connector.Error as e:
        if conexion:
            conexion.rollback()
        messagebox.showerror("❌Error", f"Error al agregar autorización: {e}")

    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()

def cargar_contratos():
    conexion = conectar_bd()
    cursor = conexion.cursor()
    cursor.execute("SELECT id_contrato, contrato FROM contratos")
    contratos = [f"{id} - {nombre}" for id, nombre in cursor.fetchall()]
    cursor.close()
    conexion.close()
    return contratos

def asignar_importes_a_contratos(ventana_padre, contratos_ids_nombres, monto_total):
    importes = {}
    confirmacion = {"aceptado": False}

    ventana_importes = tk.Toplevel(ventana_padre)
    ventana_importes.title("Asignar porcentajes a contratos")
    centrar_ventana(ventana_importes, 450, 350)
    tk.Label(ventana_importes, text="Asigne el porcentaje de pago correspondiente para cada contrato", font=("Arial", 10, "bold")).pack(pady=10)

    entries = {}

    for id_contrato, nombre_contrato in contratos_ids_nombres:
        frame = tk.Frame(ventana_importes)
        frame.pack(pady=5)
        tk.Label(frame, text=nombre_contrato, width=30, anchor="w").pack(side="left")
        entry = tk.Entry(frame, width=10)
        entry.pack(side="left", padx=10)
        entries[id_contrato] = entry

    def confirmar():
        total_porcentaje = 0
        for id_contrato, entry in entries.items():
            valor = entry.get()
            if not valor:
                messagebox.showwarning("Campo vacío", f"Falta el porcentaje del contrato {id_contrato}")
                return
            try:
                porcentaje = float(valor)
                if porcentaje < 0 or porcentaje > 100:
                    raise ValueError
                total_porcentaje += porcentaje
                importe = round(monto_total * (porcentaje / 100), 2)
                importes[id_contrato] = importe
            except ValueError:
                messagebox.showwarning("Error", f"Porcentaje inválido para contrato {id_contrato}")
                return

        if round(total_porcentaje, 2) != 100.00:
            messagebox.showwarning("Error", f"La suma de los porcentajes debe ser exactamente 100%.\nActual: {total_porcentaje}%")
            return

        confirmacion["aceptado"] = True
        ventana_importes.destroy()

    def cancelar():
        ventana_importes.destroy()

    boton_frame = tk.Frame(ventana_importes)
    boton_frame.pack(pady=10)
    tk.Button(boton_frame, text="Aceptar", command=confirmar, bg="#006600", fg="white").pack(side="left", padx=10)
    tk.Button(boton_frame, text="Cancelar", command=cancelar, bg="red", fg="white").pack(side="left", padx=10)

    ventana_importes.grab_set()
    ventana_padre.wait_window(ventana_importes)

    if confirmacion["aceptado"]:
        return importes
    else:
        return None

def modificar_importes_contratos(ventana_padre, id_autorizacion):
    importes = {}
    confirmacion = {"aceptado": False}

    # Obtener contratos actuales desde la base de datos
    conexion = conectar_bd()
    cursor = conexion.cursor()
    cursor.execute("""
        SELECT ac.id_contrato, c.contrato, ac.importe
        FROM autorizacion_contratos ac
        JOIN contratos c ON ac.id_contrato = c.id_contrato
        WHERE ac.id_autorizacion = %s
    """, (id_autorizacion,))
    contratos = cursor.fetchall()
    print("🧩 Contratos recuperados desde la base de datos:", contratos)
    print("🔎 Ejecutando consulta para ID:", id_autorizacion)


    cursor.close()
    conexion.close()

    if not contratos:
        messagebox.showinfo("Sin contratos", "No hay contratos asociados a esta autorización.")
        return None

    # Crear ventana para modificar los importes
    ventana = tk.Toplevel(ventana_padre)
    ventana.title("Modificar importes de contratos")
    centrar_ventana(ventana, 400, 300)
    tk.Label(ventana, text="Modifica el importe de cada contrato:", font=("Arial", 12, "bold")).pack(pady=10)

    entries = {}

    for id_contrato, nombre, importe in contratos:
        frame = tk.Frame(ventana)
        frame.pack(pady=5, fill="x", padx=20)
        tk.Label(frame, text=nombre, width=30, anchor="w").pack(side="left")
        entry = tk.Entry(frame, width=15)
        entry.insert(0, str(importe))
        entry.pack(side="left")
        entries[id_contrato] = entry

    def confirmar():
        for id_contrato, entry in entries.items():
            valor = entry.get()
            if not valor:
                messagebox.showwarning("Campo vacío", f"Falta el importe del contrato {id_contrato}")
                return
            try:
                importes[id_contrato] = float(valor)
            except ValueError:
                messagebox.showwarning("Error", f"Importe inválido: {valor}")
                return
        confirmacion["aceptado"] = True
        ventana.destroy()

    def cancelar():
        ventana.destroy()

    frame_botones = tk.Frame(ventana)
    frame_botones.pack(pady=10)
    tk.Button(frame_botones, text="Aceptar", command=confirmar, bg="green", fg="white").pack(side="left", padx=10)
    tk.Button(frame_botones, text="Cancelar", command=cancelar, bg="red", fg="white").pack(side="left", padx=10)

    ventana.grab_set()
    ventana_padre.wait_window(ventana)

    if confirmacion["aceptado"]:
        return importes
    else:
        return None

#Funcion para agregar los articulos comprados
articulos_lista = []
def agregar_articulo(entry_cantidad, entry_unidad, entry_articulo, entry_observaciones, tree):
    cantidad = entry_cantidad.get()
    unidad = entry_unidad.get()
    articulo = entry_articulo.get("1.0", "end-1c") 
    observaciones = entry_observaciones.get("1.0", "end-1c")

    if not (cantidad and unidad and articulo):
        messagebox.showwarning("Campos vacíos", "Debe ingresar cantidad, unidad y artículo.")
        return
    
    articulos_lista.append((cantidad, unidad, articulo, observaciones))
    
    tree.insert("", "end", values=(cantidad, unidad, articulo, observaciones))

    # Limpiar campos
    entry_cantidad.delete(0, tk.END)
    entry_unidad.delete(0, tk.END)
    entry_articulo.delete("1.0", "end")
    entry_observaciones.delete("1.0", "end")
        
#Carga las autorizaciones y las muestra en la tabla
def cargar_autorizaciones(tree):

    for row in tree.get_children():
        tree.delete(row)  

    conexion = None
    cursor = None

    try:
        #Conectar a la base de datos
        conexion = conectar_bd()
        if conexion is None:
            print ("❌No se pudo establecer la conexion")
            return
        cursor = conexion.cursor()

        #Se ejecuta la consulta
        query = "SELECT id_autorizacion, tipo_solicitud, solicitante, monto, fecha_limite_pago , fecha_solicitud, estado FROM autorizacionescompra WHERE estado <> 'Autorizado'"
        cursor.execute(query)
        autorizaciones = cursor.fetchall()

        #Muestra resultados
        for autorizacion in autorizaciones:
            tree.insert("", "end", values=autorizacion)

    except mysql.connector.Error as e:
        print(f"❌Error al cargar autorizaciones: {e}")

    finally:
        #Cierra el cursor y la conexion si fueron creados correctamente
        if cursor is not None:
            cursor.close()
        if conexion is not None:
            conexion.close()

#Carga los articulos y los muestra en la tabla principal
def cargar_articulos(tree):

    for row in tree.get_children():
        tree.delete(row)  

        
    conexion = None
    cursor = None

    try:
        #Conectar a la base de datos
        conexion = conectar_bd()
        if conexion is None:
            print ("❌No se pudo establecer la conexion")
            return
        cursor = conexion.cursor()

        #Se ejecuta la consulta
        query = "SELECT cantidad, unidad, articulo , observaciones FROM articulosautorizacion"
        cursor.execute(query)
        articulos = cursor.fetchall()

        #Muestra resultados
        for articulos in articulos:
            tree.insert("", "end", values=articulos)

    except mysql.connector.Error as e:
        print(f"❌Error al cargar articulos: {e}")

    finally:
        #Cierra el cursor y la conexion si fueron creados correctamente
        if cursor is not None:
            cursor.close()
        if conexion is not None:
            conexion.close()

def limpiar_formulario(entry_consecutivo, combo_tipo, entry_solicitante, entry_puesto, entry_area, entry_fecha_solicitud, 
                       entry_fecha_requerida, entry_monto, combo_proveedor,
                       combo_instruccion, entry_flimite, listbox_contratos, entry_IVA, entry_subtotal):

    entry_consecutivo.delete(0, tk.END)
    combo_tipo.set("")  
    entry_solicitante.delete(0, tk.END)
    entry_puesto.delete(0, tk.END)
    entry_area.delete(0, tk.END)
    entry_fecha_solicitud.delete(0, tk.END)
    entry_fecha_requerida.delete(0, tk.END)
    entry_monto.delete(0, tk.END)
    combo_proveedor.set("")
    combo_instruccion.set("")
    entry_flimite.delete(0, tk.END)
    listbox_contratos.selection_clear(0, tk.END)
    entry_IVA.delete(0, tk.END)
    entry_subtotal.delete(0, tk.END)

#Funcion para limpiar la tabla de articulos
def limpiar_tabla(tree):
    #Elimina todos los registros de la tabla de artículos en la interfaz
    for row in tree.get_children():
        tree.delete(row)
    articulos_lista.clear()

def Modificar_autorizacion(
    m_fecha_solicitud, m_fecha_requerida, m_flimite, m_proveedor,
    m_subtotal, m_IVA, m_monto, m_consecutivo,
    articulos_lista, contratos_ids_nombres, ventana_padre):

    conexion = None
    cursor = None
    try:
        conexion = conectar_bd()
        cursor = conexion.cursor()

        # Obtener datos actuales
        cursor.execute("""
            SELECT fecha_solicitud, fecha_requerida, fecha_limite_pago, id_proveedor, subtotal, IVA, monto
            FROM autorizacionescompra
            WHERE id_autorizacion = %s
        """, (m_consecutivo,))
        datos_actuales_aut = cursor.fetchone()
        print("📌 ID que llega para modificar contratos:", m_consecutivo)


        cursor.execute("""
            SELECT cantidad, unidad, articulo, observaciones
            FROM articulosautorizacion
            WHERE id_autorizacion = %s
        """, (m_consecutivo,))
        articulos_actuales = cursor.fetchall()

        nuevos_datos = (m_fecha_solicitud, m_fecha_requerida, m_flimite, m_proveedor, m_subtotal, m_IVA, m_monto)
        if nuevos_datos == datos_actuales_aut and articulos_actuales == articulos_lista:
            messagebox.showinfo("Sin cambios", "No se detectaron cambios en la autorización.")
            return

        # Actualizar autorización
        cursor.execute("""
            UPDATE autorizacionescompra
            SET fecha_solicitud = %s, fecha_requerida = %s, fecha_limite_pago = %s,
                id_proveedor = %s, subtotal = %s, IVA = %s, monto = %s, estado = 'Pendiente'
            WHERE id_autorizacion = %s
        """, (m_fecha_solicitud, m_fecha_requerida, m_flimite, m_proveedor,
              m_subtotal, m_IVA, m_monto, m_consecutivo))
        
        # Obtener nuevos importes por contrato desde ventana
        importes_contratos = modificar_importes_contratos(ventana_padre, m_consecutivo)
        if importes_contratos is None:
            messagebox.showwarning("Cancelado", "Modificación cancelada.")
            return

        suma_importes = round(sum(importes_contratos.values()), 2)
        if round(m_monto, 2) != suma_importes:
            messagebox.showerror("Error", f"Suma de importes ({suma_importes}) no coincide con el monto total ({m_monto})")
            return

        # Borrar artículos y contratos anteriores
        cursor.execute("DELETE FROM articulosautorizacion WHERE id_autorizacion = %s", (m_consecutivo,))
        cursor.execute("DELETE FROM autorizacion_contratos WHERE id_autorizacion = %s", (m_consecutivo,))
        cursor.execute("DELETE FROM solicitud_contratos WHERE id_solicitud = (SELECT id_solicitud FROM solicitudespago WHERE id_autorizacion = %s)", (m_consecutivo,))

        # Insertar nuevos artículos
        for articulo in articulos_lista:
            cantidad, unidad, nombre_articulo, observaciones = articulo
            cursor.execute("""
                INSERT INTO articulosautorizacion (id_autorizacion, cantidad, unidad, articulo, observaciones)
                VALUES (%s, %s, %s, %s, %s)
            """, (m_consecutivo, cantidad, unidad, nombre_articulo, observaciones))

        id_solicitud = None
        cursor.execute("SELECT id_solicitud FROM solicitudespago WHERE id_autorizacion = %s", (m_consecutivo,))
        resultado = cursor.fetchone()
        if resultado:
            id_solicitud = resultado[0]

        for id_contrato, importe in importes_contratos.items():
            cursor.execute("""
                INSERT INTO autorizacion_contratos (id_autorizacion, id_contrato, importe)
                VALUES (%s, %s, %s)
            """, (m_consecutivo, id_contrato, importe))

            if id_solicitud:
                cursor.execute("""
                    INSERT INTO solicitud_contratos (id_solicitud, id_contrato, importe)
                    VALUES (%s, %s, %s)
                """, (id_solicitud, id_contrato, importe))

        # Actualizar solicitud de pago
        cursor.execute("""
            UPDATE solicitudespago
            SET fecha_solicitud = %s, fecha_limite_pago = %s, id_proveedor = %s,
                importe = %s, IVA = %s, SUBTOTAL = %s, estado = 'Pendiente'
            WHERE id_autorizacion = %s
        """, (m_fecha_solicitud, m_flimite, m_proveedor, m_monto, m_IVA, m_subtotal, m_consecutivo))

        conexion.commit()
        generar_excel_modificado()
        mSolicitud_excel(m_consecutivo)
        messagebox.showinfo("✅ Éxito", "Autorización, artículos, contratos y solicitud de pago modificados.")

    except mysql.connector.Error as e:
        if conexion:
            conexion.rollback()
        messagebox.showerror("❌ Error", f"Error al modificar: {e}")
    finally:
        if cursor: cursor.close()
        if conexion: conexion.close()

def ventana_modificar(tree):
    ventana = tk.Toplevel()
    ventana.title("Modificar Autorización")
    centrar_ventana(ventana, 1000, 600)

    def agregar_articulo_local():
        cantidad = entry_cantidad.get()
        unidad = entry_unidad.get()
        articulo = entry_articulo.get("1.0", "end-1c") 
        observaciones = entry_observaciones.get("1.0", "end-1c")

        if not (cantidad and unidad and articulo):
            messagebox.showwarning("Campos vacíos", "Debe ingresar cantidad, unidad y artículo.")
            return

        tree_articulos.insert("", "end", values=(cantidad, unidad, articulo, observaciones))

        entry_cantidad.delete(0, tk.END)
        entry_unidad.delete(0, tk.END)
        entry_articulo.delete("1.0", "end")
        entry_observaciones.delete("1.0", "end")

    def eliminar_articulo():
        seleccion = tree_articulos.selection()
        if not seleccion:
            messagebox.showwarning("Eliminar artículo", "Selecciona un artículo para eliminar.")
            return
        tree_articulos.delete(seleccion[0])

    seleccion = tree.selection()
    if not seleccion:
        messagebox.showwarning("Selecciona una autorización", "Por favor selecciona una autorización para modificar.")
        return

    item = tree.item(seleccion)
    valores = item["values"]
    id_autorizacion = valores[0]
    estado = valores[6]  
    if estado != "Modificar":
        messagebox.showwarning("Atención", f"No se puede modificar esta autorización.\nEstado actual: {estado}")
        return

    conexion = conectar_bd()
    cursor = conexion.cursor()
    cursor.execute("""
        SELECT fecha_solicitud, fecha_requerida, fecha_limite_pago, id_proveedor, subtotal, IVA, monto
        FROM autorizacionescompra
        WHERE id_autorizacion = %s
    """, (id_autorizacion,))
    datos_aut = cursor.fetchone()

    cursor.execute("""
        SELECT cantidad, unidad, articulo, observaciones
        FROM articulosautorizacion
        WHERE id_autorizacion = %s
    """, (id_autorizacion,))
    articulos = cursor.fetchall()

    cursor.execute("""
        SELECT c.id_contrato, c.contrato
        FROM autorizacion_contratos ac
        JOIN contratos c ON ac.id_contrato = c.id_contrato
        WHERE ac.id_autorizacion = %s
    """, (id_autorizacion,))
    contratos = cursor.fetchall()
    contratos_ids_nombres = [(str(c[0]), f"{c[0]} - {c[1]}") for c in contratos]

    conexion.close()

    campos = ["Fecha solicitud", "Fecha requerida", "Fecha límite", "Proveedor", "Subtotal", "IVA", "Monto"]
    entradas = []

    for i, campo in enumerate(campos):
        rel_y = 0.05 + i * 0.05
        tk.Label(ventana, text=campo).place(relx=0.05, rely=rel_y)

        if campo == "Proveedor":
            combo = ttk.Combobox(ventana, values=cargar_proveedores(), width=27, state="readonly")
            id_actual = datos_aut[i]
            nombre_actual = ""
            for proveedor in cargar_proveedores():
                id_str, nombre = proveedor.split(" - ", 1)
                if int(id_str) == id_actual:
                    nombre_actual = proveedor
                    break
            combo.set(nombre_actual)
            combo.place(relx=0.4, rely=rel_y)
            entradas.append(combo)
        else:
            entry = tk.Entry(ventana, width=30)
            entry.insert(0, datos_aut[i])
            entry.place(relx=0.4, rely=rel_y)
            entradas.append(entry)

    columnas = ("Cantidad", "Unidad", "Artículo", "Observaciones")
    tree_articulos = ttk.Treeview(ventana, columns=columnas, show="headings", height=6)
    for col in columnas:
        tree_articulos.heading(col, text=col)
        if col == "Observaciones":
            tree_articulos.column(col, width=400, anchor="w")
        else:
            tree_articulos.column(col, width=100, anchor="center")
    tree_articulos.place(relx=0.05, rely=0.40)

    for articulo in articulos:
        tree_articulos.insert("", tk.END, values=articulo)

    tk.Label(ventana, text="¿Necesitas agregar más artículos?", fg="#990000").place(relx=0.25, rely=0.69)
    tk.Label(ventana, text="Cantidad").place(relx=0.05, rely=0.75)
    entry_cantidad = ttk.Entry(ventana, width=10)
    entry_cantidad.place(relx=0.11, rely=0.75)

    tk.Label(ventana, text="Unidad").place(relx=0.27, rely=0.75)
    entry_unidad = ttk.Entry(ventana, width=10)
    entry_unidad.place(relx=0.33, rely=0.75)

    tk.Label(ventana, text="Artículo").place(relx=0.49, rely=0.75)
    entry_articulo = tk.Text(ventana, height=2, width=20, font=("Arial", 10))
    entry_articulo.place(relx=0.55, rely=0.75)

    tk.Label(ventana, text="Observaciones").place(relx=0.05, rely=0.82)
    entry_observaciones = tk.Text(ventana, height=3, width=50, font=("Arial", 10))
    entry_observaciones.place(relx=0.22, rely=0.82)

    def editar_articulo():
        item_sel = tree_articulos.selection()
        if not item_sel:
            messagebox.showwarning("Selecciona", "Selecciona un artículo para editar.")
            return

        datos = tree_articulos.item(item_sel, "values")
        top_edit = tk.Toplevel(ventana)
        top_edit.title("Editar artículo")
        centrar_ventana(top_edit, 400, 300)
        labels = ["Cantidad", "Unidad", "Artículo", "Observaciones"]
        entradas_local = []

        for i, (label_text, valor) in enumerate(zip(labels, datos)):
            tk.Label(top_edit, text=label_text).pack()
            e = tk.Entry(top_edit)
            e.insert(0, valor)
            e.pack()
            entradas_local.append(e)

        def guardar_edicion():
            nuevos = [e.get() for e in entradas_local]
            tree_articulos.item(item_sel, values=nuevos)
            top_edit.destroy()

        tk.Button(top_edit, text="Guardar", command=guardar_edicion).pack(pady=10)

    def guardar():
        datos_aut = [e.get() for e in entradas]
        try:
            id_proveedor = int(datos_aut[3].split(" - ")[0])
        except (IndexError, ValueError):
            messagebox.showerror("Error", "Formato de proveedor inválido. Debes seleccionar uno de la lista.")
            return
        datos_aut[3] = id_proveedor

        try:
            datos_aut[4] = float(datos_aut[4])  # Subtotal
            datos_aut[5] = float(datos_aut[5])  # IVA
            datos_aut[6] = float(datos_aut[6])  # Monto
        except ValueError:
            messagebox.showerror("Error", "Los campos 'Subtotal', 'IVA' y 'Monto' deben ser numéricos.")
            return


        articulos_lista = []
        for item in tree_articulos.get_children():
            valores = tree_articulos.item(item)["values"]
            if len(valores) != 4:
                messagebox.showwarning("Artículo incompleto", "Llene todos los campos.")
                return
            articulos_lista.append(valores)

        if not articulos_lista:
            messagebox.showwarning("Sin artículos", "Debes agregar al menos un artículo.")
            return

        # Obtener contratos relacionados
        conexion = conectar_bd()
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT ac.id_contrato, c.contrato
            FROM autorizacion_contratos ac
            JOIN contratos c ON ac.id_contrato = c.id_contrato
            WHERE ac.id_autorizacion = %s
        """, (id_autorizacion,))
        contratos_relacionados = cursor.fetchall()
        cursor.close()
        conexion.close()
        print (contratos_relacionados)
        Modificar_autorizacion(*datos_aut, id_autorizacion, articulos_lista, contratos_relacionados, ventana)
        ventana.destroy()
        cargar_autorizaciones(tree)

    tk.Button(ventana, text="➕ Agregar artículo", command=agregar_articulo_local, bg="lightblue").place(relx=0.77, rely=0.75)
    tk.Button(ventana, text="❌ Eliminar artículo", command=eliminar_articulo, bg="orange").place(relx=0.77, rely=0.45)
    tk.Button(ventana, text="✏️ Editar artículo", command=editar_articulo).place(relx=0.77, rely=0.52)

    tk.Button(ventana, text="✅ Guardar cambios", command=guardar, bg="green", font=("Arial", 10, "bold")).place(relx=0.4, rely=0.93)
    tk.Button(ventana, text="Cancelar", command=ventana.destroy, bg="red", font=("Arial", 10, "bold")).place(relx=0.2, rely=0.93)

#Funcion para generar el excel
def generar_excel(entry_consecutivo, combo_tipo, combo_solicitante, entry_puesto, entry_area, 
                  entry_fecha_solicitud, entry_fecha_requerida, entry_monto, 
                  combo_proveedor, combo_instruccion, articulos, tree, entry_flimite, listbox_contratos, entry_IVA, entry_subtotal, combo_moneda, entry_general, ruta_cotizacion=None):
    try:
        plantilla_path = ruta_relativa("Plantillas/Autorizaciones.xlsx")
        workbook = load_workbook(plantilla_path)
        sheet = workbook.active

        consecutivo = entry_consecutivo.get()
        tipo_solicitud = combo_tipo.get()
        solicitante = combo_solicitante.get()
        puesto = entry_puesto.get()
        area = entry_area.get()
        fecha_solicitud = entry_fecha_solicitud.get()
        fecha_requerida = entry_fecha_requerida.get()
        monto = entry_monto.get()
        proveedor = combo_proveedor.get()
        instruccion = combo_instruccion.get()
        moneda = combo_moneda.get()
        general = entry_general.get("1.0", "end").strip()


        conexion = conectar_bd()
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT c.contrato
            FROM Autorizacion_Contratos ac
            JOIN contratos c ON ac.id_contrato = c.id_contrato
            WHERE ac.id_autorizacion = %s
        """, (consecutivo,))
        contratos = [row[0] for row in cursor.fetchall()]
        texto_contratos = " / ".join(contratos)

        # Obtener ruta cotización
        cursor.execute("SELECT ruta_cotizacion FROM AutorizacionesCompra WHERE id_autorizacion = %s", (consecutivo,))
        resultado = cursor.fetchone()
        ruta_cotizacion = resultado[0] if resultado else ""
        cursor.close()
        conexion.close()
        
        # ✅ Función para escribir y recombinar celdas
        def escribir_celda(fila, columna, valor, rango_combinado=None):
            if rango_combinado:
                sheet.unmerge_cells(rango_combinado)

            celda = sheet.cell(row=fila, column=columna)

            # Convertir monto a número si corresponde
            if valor == monto:
                try:
                    valor_numerico = float(valor)
                    celda.value = valor_numerico
                except ValueError:
                    celda.value = valor  # Fallback si no se puede convertir
            else:
                celda.value = valor

            # Alinear a la izquierda si es alguno de estos campos
            if valor in (instruccion, proveedor, monto, general):
                alineacion = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                alineacion = Alignment(horizontal="center", vertical="center", wrap_text=True)

            celda.alignment = alineacion

            if rango_combinado:
                sheet.merge_cells(rango_combinado)

        # 📝 Escribir datos principales
        escribir_celda(6, 8, consecutivo)
        escribir_celda(12, 2, solicitante, "B12:D12")
        escribir_celda(39, 2, solicitante, "B39:C39")
        escribir_celda(13, 2, puesto, "B13:D13")
        escribir_celda(40, 2, puesto, "B40:C40")
        escribir_celda(14, 2, area, "B14:D14")
        escribir_celda(12, 7, fecha_solicitud, "G12:H12")
        escribir_celda(13, 7, fecha_requerida, "G13:H13")
        escribir_celda(14, 7, texto_contratos, "G14:H14")  
        escribir_celda(33, 2, monto)
        escribir_celda(32, 2, proveedor, "B32:H32")
        escribir_celda(30, 2, general, "B30:H30")  
        escribir_celda(33, 3, instruccion)

        # Concatenar instrucción con cotización y poner enlace
        nombre_archivo = os.path.basename(ruta_cotizacion)
        celda_cotizacion = sheet.cell(row=31, column=2)
        celda_cotizacion.value = nombre_archivo
        celda_cotizacion.hyperlink = ruta_cotizacion
        celda_cotizacion.style = "Hyperlink"
        celda_cotizacion.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        sheet.merge_cells("B31:H31")

        # Obtener el valor original de la celda (por ejemplo, "Monto Total")
        celda_titulo_monto = sheet.cell(row=33, column=1)  # A33
        texto_original = celda_titulo_monto.value or ""

        # Concatenar con la moneda
        celda_titulo_monto.value = f"{texto_original} ({moneda})"
        celda_titulo_monto.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


        # 🧾 Artículos desde fila 17
        fila_inicio = 17
        for i, (cantidad, unidad, articulo, observaciones) in enumerate(articulos):
            escribir_celda(fila_inicio + i, 2, cantidad)      # B
            escribir_celda(fila_inicio + i, 3, unidad)        # C
            escribir_celda(fila_inicio + i, 4, articulo)      # D
            escribir_celda(fila_inicio + i, 7, observaciones) # G



        # Combinar columnas D:F para "artículo"
        col_inicio_art = get_column_letter(4)  # D
        col_fin_art = get_column_letter(6)     # F
        sheet.merge_cells(f"{col_inicio_art}{fila_inicio}:{col_fin_art}{fila_inicio}")

        # Combinar columnas G:H para "observaciones"
        col_inicio_obs = get_column_letter(7)  # G
        col_fin_obs = get_column_letter(8)     # H
        sheet.merge_cells(f"{col_inicio_obs}{fila_inicio}:{col_fin_obs}{fila_inicio}")

        # ✅ Tipo de solicitud con marca "X"
        tipo_a_celda = {
            "Maquinaria": "B10",
            "Equipo y/o Htas": "D10",
            "Servicios": "F10",
            "Otros": "H10"
        }

        # Limpiar anteriores
        for celda in tipo_a_celda.values():
            sheet[celda].value = ""

        if tipo_solicitud in tipo_a_celda:
            celda_obj = sheet[tipo_a_celda[tipo_solicitud]]
            celda_obj.value = "X"
            celda_obj.font = Font(bold=True, color="FF0000")

        # Firma
        ruta_firma = ruta_relativa(usuario_actual["firma"])
        firma_img = ExcelImage(ruta_firma)
        firma_img.width = 160
        firma_img.height = 50
        sheet.add_image(firma_img, "B37")

        # 📁 Guardar archivo Excel
        #CARPETA_AUTORIZACIONES = ruta_relativa("Autorizaciones") #Quitar cuando se sincronicen en one drive
        CARPETA_AUTORIZACIONES = os.path.join(os.environ['USERPROFILE'], 'OneDrive - ATI', 'Documentos de AppGestor', 'Autorizaciones de Compra')
        if not os.path.exists(CARPETA_AUTORIZACIONES):
            os.makedirs(CARPETA_AUTORIZACIONES)

        output_path = os.path.join(CARPETA_AUTORIZACIONES, f"Autorizacion_{consecutivo}.xlsx")
        workbook.save(output_path)

        # 🧹 Limpiar y recargar
        limpiar_formulario(entry_consecutivo, combo_tipo, combo_solicitante, entry_puesto, entry_area, 
                        entry_fecha_solicitud, entry_fecha_requerida, entry_monto, 
                        combo_proveedor, combo_instruccion, entry_flimite, listbox_contratos, entry_IVA, entry_subtotal)
        cargar_autorizaciones(tree)
        articulos_lista.clear()
        limpiar_tabla(tree)

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo generar el archivo Excel: {e}")
        return
    # 📄 Convertir a PDF
    ruta_pdf = output_path.replace(".xlsx", ".pdf")
    convertir_excel_a_pdf(output_path, ruta_pdf)

def generar_excel_modificado(consecutivo, solicitante, puesto, area, fecha_solicitud, fecha_requerida, monto, proveedor, instruccion, tipo_solicitud, articulos):
    try:
        plantilla_path = ruta_relativa("Plantillas/Autorizaciones.xlsx")
        workbook = load_workbook(plantilla_path)
        sheet = workbook.active

        # Escribir campos principales
        def escribir_celda(fila, columna, valor, rango_combinado=None):
            if rango_combinado:
                sheet.unmerge_cells(rango_combinado)
            celda = sheet.cell(row=fila, column=columna)
            celda.value = valor
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text= True)
            if rango_combinado:
                sheet.merge_cells(rango_combinado)

        conexion = conectar_bd()
        cursor = conexion.cursor()
        cursor.execute("SELECT c.contrato FROM Autorizacion_Contratos ac JOIN contratos c ON ac.id_contrato = c.id_contrato WHERE ac.id_autorizacion = %s", (consecutivo,))
        contratos = [row[0] for row in cursor.fetchall()]
        texto_contratos = " / ".join(contratos)

        cursor.execute("SELECT ruta_cotizacion FROM AutorizacionesCompra WHERE id_autorizacion = %s", (consecutivo,))
        resultado = cursor.fetchone()
        ruta_cotizacion = resultado[0] if resultado else ""
        cursor.close()
        conexion.close()

        escribir_celda(6, 8, consecutivo)
        escribir_celda(12, 2, solicitante, "B12:D12")
        escribir_celda(39, 2, solicitante, "B39:C39")
        escribir_celda(13, 2, puesto, "B13:D13")
        escribir_celda(40, 2, puesto, "B40:C40")
        escribir_celda(14, 2, area, "B14:D14")
        escribir_celda(12, 7, fecha_solicitud, "G12:H12")
        escribir_celda(13, 7, fecha_requerida, "G13:H13")
        escribir_celda(14, 7, texto_contratos, "G14:H14")
        escribir_celda(33, 2, monto, "B33:H33")
        escribir_celda(32, 2, proveedor, "B32:H32")
        escribir_celda(30, 2, instruccion, "B30:H30")

        celda_cotizacion = sheet.cell(row=31, column=2)
        celda_cotizacion.value = os.path.basename(ruta_cotizacion)
        celda_cotizacion.hyperlink = ruta_cotizacion
        celda_cotizacion.style = "Hyperlink"
        celda_cotizacion.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        sheet.merge_cells("B31:H31")

        fila_inicio = 17
        for i, (cantidad, unidad, articulo, observaciones) in enumerate(articulos):
            escribir_celda(fila_inicio + i, 2, cantidad)
            escribir_celda(fila_inicio + i, 3, unidad)
            escribir_celda(fila_inicio + i, 4, articulo)
            escribir_celda(fila_inicio + i, 7, observaciones)
            sheet.merge_cells(f"D{fila_inicio + i}:F{fila_inicio + i}")
            sheet.merge_cells(f"G{fila_inicio + i}:H{fila_inicio + i}")

        tipo_a_celda = {
            "Maquinaria": "B10",
            "Equipo y/o Htas": "D10",
            "Servicios": "F10",
            "Otros": "H10"
        }
        for celda in tipo_a_celda.values():
            sheet[celda].value = ""
        if tipo_solicitud in tipo_a_celda:
            sheet[tipo_a_celda[tipo_solicitud]].value = "X"
            sheet[tipo_a_celda[tipo_solicitud]].font = Font(bold=True, color="FF0000")

        ruta_firma = ruta_relativa(usuario_actual["firma"])
        firma_img = ExcelImage(ruta_firma)
        firma_img.width = 120
        firma_img.height = 50
        sheet.add_image(firma_img, "B37")

        carpeta = os.path.join(os.environ['USERPROFILE'], 'OneDrive - ATI', 'Documentos de AppGestor', 'Autorizaciones de Compra')
        if not os.path.exists(carpeta):
            os.makedirs(carpeta)

        output_path = os.path.join(carpeta, f"Autorizacion_{consecutivo}.xlsx")
        workbook.save(output_path)


        # Convertir a PDF
        ruta_pdf = output_path.replace(".xlsx", ".pdf")
        convertir_excel_a_pdf(output_path, ruta_pdf)

    except Exception as e:
        messagebox.showerror("Error al regenerar Excel", f"{e}")

#Interfaz de Usuario
def hex_a_rgb(hex_color):
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2 ,4))

def crear_degradado_vertical(canvas, ancho, alto, color_inicio, color_fin):
    canvas.delete("degradado")

    rgb_inicio = hex_a_rgb(color_inicio)
    rgb_fin = hex_a_rgb(color_fin)

    pasos = alto // 2
    for i in range(pasos):
        y = alto - i - 1
        r = int(rgb_inicio[0] + (rgb_fin[0] - rgb_inicio[0]) * i / pasos)
        g = int(rgb_inicio[1] + (rgb_fin[1] - rgb_inicio[1]) * i / pasos)
        b = int(rgb_inicio[2] + (rgb_fin[2] - rgb_inicio[2]) * i / pasos)
        color = f"#{r:02x}{g:02x}{b:02x}"
        canvas.create_line(0, y, ancho, y, fill=color, tags="degradado")

    canvas.create_rectangle(0, 0, ancho, alto // 2, fill=color_fin, outline="", tags="degradado")

def gestionar_autorizaciones(rol, volver_menu_callback):

    ventana = tk.Tk()
    ventana.title("Gestión de Autorizaciones de Compra")
    centrar_ventana(ventana, 1200, 600)

    canvas = tk.Canvas(ventana)
    canvas.pack(fill="both", expand=True)


    def actualizar_degradado(event):
        # Obtener las dimensiones del canvas
        ancho = canvas.winfo_width()
        alto = canvas.winfo_height()
        crear_degradado_vertical(canvas, ancho, alto, "#8B0000", "#FFFFFF")

    # Actualizar el fondo degradado al cambiar el tamaño de la ventana
    canvas.bind("<Configure>", actualizar_degradado)

    # Inicializar el degradado en el tamaño actual de la ventana
    ventana.after(100, lambda: actualizar_degradado(None))

    # Función para calcular posiciones relativas
    def pos(x, y):
        return {"relx": x, "rely": y, "anchor": "w"}   
    
    #Etiqueta de instruccion
    tk.Label(ventana, text="Datos Generales de la Autorizacion de Compra", font=("Arial", 12, "bold"), bg="white").place(**pos(0.07, 0.03))
    
    tk.Label(ventana, text="Consecutivo: ", font=("Arial", 10, "bold"), bg="white").place(**pos(0.05, 0.10))
    entry_consecutivo = ttk.Entry(ventana)
    entry_consecutivo.place(**pos(0.3, 0.10))


    tk.Label(ventana, text="Tipo de Solicitud: ", font=("Arial", 10, "bold"), bg="white").place(**pos(0.05, 0.15))
    combo_tipo = ttk.Combobox(ventana, values=["Maquinaria", "Equipo y/o Htas", "Servicios", "Otros"])
    combo_tipo.place(**pos(0.3, 0.15))

    
    tk.Label(ventana, text="Solicitante:", font=("Arial", 10, "bold"), bg="white").place(**pos(0.05, 0.20))
    entry_solicitante = ttk.Entry(ventana, state="readonly")
    entry_solicitante.place(**pos(0.3, 0.20), relwidth=0.20)
    entry_solicitante.config(state="normal")
    entry_solicitante.insert(0, usuario_actual["nombre"])
    entry_solicitante.config(state="readonly")

    tk.Label(ventana, text="Puesto:", font=("Arial", 10, "bold"), bg="white").place(**pos(0.05, 0.25))
    entry_puesto = ttk.Entry(ventana, state="readonly")
    entry_puesto.place(**pos(0.3, 0.25), relwidth=0.20)
    entry_puesto.config(state="normal")
    entry_puesto.insert(0, usuario_actual["puesto"])
    entry_puesto.config(state="readonly")

    tk.Label(ventana, text="Área:", font=("Arial", 10, "bold"), bg="white").place(**pos(0.05, 0.30))
    entry_area = ttk.Entry(ventana)
    entry_area.place(**pos(0.3, 0.30))

    tk.Label(ventana, text="Fecha de Solicitud (AAAA/MM/DD):", font=("Arial", 10, "bold"), bg="white").place(**pos(0.05, 0.35))
    entry_fecha_solicitud = DateEntry(ventana, date_pattern='yyyy/mm/dd')
    entry_fecha_solicitud.place(**pos(0.3, 0.35))

    tk.Label(ventana, text="Fecha Requerida (AAAA/MM/DD):", font=("Arial", 10, "bold"), bg="white").place(**pos(0.05, 0.40))
    entry_fecha_requerida = DateEntry(ventana, date_pattern='yyyy/mm/dd')
    entry_fecha_requerida.place(**pos(0.3, 0.40))

    tk.Label(ventana, text="Proyecto(s) / Contrato(s):", font=("Arial", 10, "bold"), bg="white").place(**pos(0.05, 0.5))
    frame_contratos = ttk.Frame(ventana, relief="solid", borderwidth=1)
    frame_contratos.place(**pos(0.3, 0.5), relwidth=0.2, relheight=0.15)
    scrollbar_y = ttk.Scrollbar(frame_contratos, orient="vertical")
    scrollbar_y.pack(side="right", fill="y")
    listbox_contratos = tk.Listbox(frame_contratos, selectmode="multiple", height=5, font=("Arial", 10),
                                yscrollcommand=scrollbar_y.set, exportselection=False)
    listbox_contratos.pack(side="left", fill="both", expand=True)

    scrollbar_y.config(command=listbox_contratos.yview)

    for contrato in cargar_contratos():
        listbox_contratos.insert(tk.END, contrato)

    tk.Label(ventana, text="Limite de Pago:", font=("Arial", 10, "bold"), bg="#ffebeb").place(**pos(0.05, 0.60))
    entry_flimite = ttk.Entry(ventana)
    entry_flimite.place(**pos(0.3, 0.60))

        # --- Proveedor ---
    tk.Label(ventana, text="Proveedor:", font=("Arial", 10, "bold"), bg="#ffebeb").place(relx=0.05, rely=0.65)

    combo_proveedor = ttk.Combobox(ventana)
    combo_proveedor.place(**pos(0.25, 0.66), relwidth=0.20)

    proveedores_originales = []

    def actualizar_combobox_proveedores():
        nonlocal proveedores_originales
        proveedores = cargar_proveedores()
        if proveedores:
            proveedores_originales = proveedores
            combo_proveedor['values'] = proveedores_originales

    def filtrar_proveedores(event):
        texto = combo_proveedor.get().lower()
        if not texto:
            combo_proveedor['values'] = proveedores_originales
        else:
            combo_proveedor['values'] = [p for p in proveedores_originales if texto in p.lower()]

    combo_proveedor.bind("<KeyRelease>", filtrar_proveedores)

    tk.Button(ventana, text="Actualizar", command=actualizar_combobox_proveedores, font=("Arial", 10, "bold")).place(relx=0.47, rely=0.64)

    actualizar_combobox_proveedores()

    #Etiqueta de instruccion de llenado de articulos
    tk.Label(ventana, text="Ingrese los datos de la compra", font=("Arial", 12, "bold"), bg="white").place(**pos(0.57, 0.03))

    tk.Label(ventana, text="Cantidad:", font=("Arial", 10, "bold"), bg="white").place(**pos(0.55, 0.10))
    entry_cantidad = ttk.Entry(ventana)
    entry_cantidad.place(**pos(0.75, 0.10))

    tk.Label(ventana, text="Unidad:", font=("Arial", 10, "bold"), bg="white").place(**pos(0.55, 0.15))
    entry_unidad = ttk.Entry(ventana)
    entry_unidad.place(**pos(0.75, 0.15))

    tk.Label(ventana, text="Descripción:", font=("Arial", 10, "bold"), bg="white").place(**pos(0.55, 0.20))
    entry_articulo = tk.Text(ventana, wrap="word", font=("Arial", 10))
    entry_articulo.place(relx=0.55, rely=0.23, relwidth=0.16, relheight=0.08)

    tk.Label(ventana, text="Observación:", font=("Arial", 10, "bold"), bg="white").place(**pos(0.75, 0.20))
    entry_observaciones = tk.Text(ventana, wrap="word", font=("Arial", 10))
    entry_observaciones.place(relx=0.75, rely=0.23, relwidth=0.18, relheight=0.08)

    tk.Label(ventana, text="Observaciones Generales:", font=("Arial", 10, "bold"), bg="white").place(relx=0.55, rely=0.38)
    entry_general = tk.Text(ventana, wrap="word", font=("Arial", 10))
    entry_general.place(relx=0.55, rely=0.41, relwidth=0.38, relheight=0.08)

    #Informacion de costos
    tk.Label(ventana, text="Ingrese los datos del pago de toda la compra", font=("Arial", 9, "bold"), bg="#ffebeb").place(**pos(0.66,0.56))
    tk.Label(ventana, text="Subtotal: ", font=("Arial", 10, "bold"), bg="#ffebeb").place(**pos(0.55, 0.60))
    entry_subtotal = ttk.Entry(ventana)
    entry_subtotal.place(**pos(0.63, 0.60))

    tk.Label(ventana, text="IVA: ", font=("Arial", 10, "bold"), bg="#ffebeb").place(**pos(0.55, 0.65))
    entry_IVA = ttk.Entry(ventana)
    entry_IVA.place(**pos(0.60, 0.65))

    tk.Label(ventana, text="Monto Total: ", font=("Arial", 10, "bold"), bg="#ffebeb").place(**pos(0.75, 0.60))
    entry_monto = ttk.Entry(ventana)
    entry_monto.place(**pos(0.83, 0.60), relwidth=0.09)

    combo_moneda = ttk.Combobox(ventana, values=["MXN", "USD"], state="readonly", font=("Arial", 10))
    combo_moneda.place(**pos(0.93, 0.60),relwidth=0.06, relheight=0.04)
    combo_moneda.set("MXN")

    tk.Label(ventana, text="Forma de Pago: ", font=("Arial", 10, "bold"), bg="#ffebeb").place(**pos(0.75, 0.65))
    combo_instruccion = ttk.Combobox(ventana, values=["Transferencia Electrónica", "Tarjeta de Débito", "Efectivo"])
    combo_instruccion.place(**pos(0.85, 0.65))

    style = ttk.Style()
    style.theme_use("alt")
    style.configure("Treeview.Heading", font=("Arial", 10, "bold"), foreground="white", background="#990000")
    style.map("Treeview.Heading", background=[("!active", "#990000"), ("active", "#990000"), ("pressed", "#990000")],
          foreground=[("!active", "white"), ("active", "white"), ("pressed", "white")]) 

    # Tabla de artículos
    tree = ttk.Treeview(ventana, columns=("Cantidad", "Unidad", "Descripción", "Observaciones"), show="headings")

    tree.heading("Cantidad", text="Cantidad")
    tree.heading("Unidad", text="Unidad")
    tree.heading("Descripción", text="Descripción")
    tree.heading("Observaciones", text="Observaciones")

    # Ajustes de ancho y alineación por columna
    tree.column("Cantidad", width=80, anchor="center")
    tree.column("Unidad", width=80, anchor="center")
    tree.column("Descripción", width=300, anchor="w")
    tree.column("Observaciones", width=300, anchor="w")

    tree.place(relx=0.05, rely=0.70, relwidth=0.9, relheight=0.2)  # Tamaño relativo

    #Ventana para vizualizar las autorizaciones guardadas
    def autorizaciones(tree):
    
        ventana = tk.Toplevel()
        ventana.title("Autorizaciones Guardadas")
        centrar_ventana(ventana, 1300, 600)

        canvas = tk.Canvas(ventana)
        canvas.pack(fill="both", expand=True)

        def actualizar_degradado(event):
            # Obtener las dimensiones del canvas
            ancho = canvas.winfo_width()
            alto = canvas.winfo_height()
            crear_degradado_vertical(canvas, ancho, alto, "#8B0000", "#FFFFFF")

        ventana.after(100, lambda: actualizar_degradado(None))

        style = ttk.Style()
        style.theme_use("alt")
        style.configure("Treeview.Heading", font=("Arial", 10, "bold"), foreground="white", background="#990000")
        style.map("Treeview.Heading", background=[("!active", "#990000"), ("active", "#990000"), ("pressed", "#990000")],
            foreground=[("!active", "white"), ("active", "white"), ("pressed", "white")])

        tree = ttk.Treeview(ventana, columns=("ID","Tipo", "Solicitante", "Monto", "Fecha Limite","Fecha solicitud", "Estado"), show="headings")
        for col in ("ID","Tipo", "Solicitante", "Monto", "Fecha Limite","Fecha solicitud", "Estado"):
            tree.heading(col, text=col)
            if col == "Solicitante":
                tree.column(col, width=100, anchor="w")
            else:
                tree.column(col, width=70, anchor= "center")
        tree.place(relx=0.05, rely=0.05, relwidth=0.9, relheight=0.8)

            # 🔹 Definir colores según estado
        tree.tag_configure("Autorizado", background="#90EE90")   # Verde claro
        tree.tag_configure("Pendiente", background="#FFD580")    # Amarillo claro
        tree.tag_configure("Rechazado", background="#FF7F7F")    # Rojo claro
        tree.tag_configure("otro", background="#FFFFFF")         # Blanco

        # 🔹 Cargar datos con tags
        conexion = conectar_bd()
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT id_autorizacion, tipo_solicitud, solicitante, monto, fecha_limite_pago, fecha_solicitud, estado
            FROM autorizacionescompra
        """)
        for fila in cursor.fetchall():
            estado = fila[-1] if fila[-1] else ""
            if estado == "Autorizado":
                tree.insert("", "end", values=fila, tags=("Autorizado",))
            elif estado == "Pendiente":
                tree.insert("", "end", values=fila, tags=("Pendiente",))
            elif estado == "Rechazado":
                tree.insert("", "end", values=fila, tags=("Rechazado",))
            else:
                tree.insert("", "end", values=fila, tags=("otro",))
        
        #tk.Button(ventana, text="Modificar autorizacion", command=lambda: ventana_modificar(tree), font=("Arial", 10, "bold")).place(relx=0.8, rely=0.91, relwidth=0.15, relheight=0.05)
        tk.Button(ventana, text="Salir", command=ventana.destroy, bg="red", fg="white", font=("Arial", 10, "bold")).place(relx=0.05, rely=0.91, relwidth=0.095, relheight=0.05)

            
    tk.Button(ventana, text="Agregar Artículo", command=lambda: agregar_articulo(entry_cantidad, entry_unidad, entry_articulo, entry_observaciones, tree), foreground="white", background="#990000", font=("Arial", 10, "bold")).place(relx=0.75, rely=0.34, relwidth=0.095, relheight=0.05)

    tk.Button(ventana, text="No encuentro mi Proveedor", command=lambda: ventana_agregar_proveedor(tree=None), font=("Arial", 10,"bold")).place(relx=0.3, rely=0.91)
    tk.Button(ventana, text="Registrar Autorización", command=lambda: agregar_autorizacion(entry_consecutivo,
        combo_tipo, entry_solicitante, entry_puesto, entry_area, entry_fecha_solicitud, 
        entry_fecha_requerida, entry_monto, combo_proveedor, combo_instruccion, entry_flimite, tree, listbox_contratos, entry_IVA, entry_subtotal, combo_moneda, entry_general), font=("Arial", 9, "bold")).place(relx=0.48, rely=0.91, relwidth=0.15, relheight=0.05)

    tk.Button(ventana, text="Generar Docs", command=lambda: generar_excel(entry_consecutivo,
        combo_tipo, entry_solicitante, entry_puesto, entry_area, entry_fecha_solicitud, 
        entry_fecha_requerida, entry_monto, combo_proveedor, combo_instruccion, articulos_lista, tree, entry_flimite, listbox_contratos, entry_IVA, entry_subtotal, combo_moneda, entry_general), font=("Arial", 10, "bold")).place(relx=0.65, rely=0.91, relwidth=0.095, relheight=0.05)
    
    tk.Button(ventana,text="Autorizaciones Guardadas", command=lambda: autorizaciones(tree), font=("Arial", 10, "bold")).place(relx=0.78, rely=0.91, relwidth=0.15, relheight=0.05)

    tk.Button(ventana, text="Salir", command= lambda: salir(volver_menu_callback, ventana), bg="red", fg="white", font=("Arial", 10, "bold")).place(relx=0.05, rely=0.91, relwidth=0.095, relheight=0.05)

    ventana.mainloop()