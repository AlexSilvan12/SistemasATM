
import tkinter as tk
import os
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
from database import conectar_bd
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from utils import centrar_ventana, ruta_relativa

def obtener_gastos_por_contrato(nombre_contrato=None, mes=None, año=None):
    conexion = conectar_bd()
    cursor = conexion.cursor()

    query = """
        SELECT 
            sp.id_solicitud,
            ac.tipo_solicitud,
            sc.importe,
            sp.fecha_solicitud,
            sp.concepto,
            p.nombre AS proveedor,
            sp.subtotal,
            sp.IVA,
            sp.importe as total         
        FROM 
            solicitudespago sp
        JOIN 
            autorizacionescompra ac ON sp.id_autorizacion = ac.id_autorizacion
        JOIN 
            solicitud_contratos sc ON sp.id_solicitud = sc.id_solicitud
        JOIN 
            contratos c ON sc.id_contrato = c.id_contrato
        JOIN
            proveedores p ON sp.id_proveedor = p.id_proveedor
        WHERE 
            (%s IS NULL OR c.contrato = %s)
            AND (%s IS NULL OR MONTH(sp.fecha_solicitud) = %s)
            AND (%s IS NULL OR YEAR(sp.fecha_solicitud) = %s)
    """
    params = (nombre_contrato, nombre_contrato, mes, mes, año, año)
    cursor.execute(query, params)
    datos = cursor.fetchall()
    cursor.close()
    conexion.close()
    return datos

def cargar_contratos(combo):
    conexion = conectar_bd()
    cursor = conexion.cursor()
    cursor.execute("SELECT contrato FROM contratos")
    contratos = [row[0] for row in cursor.fetchall()]
    cursor.close()
    conexion.close()

    opciones = ["Todos los contratos"] + contratos
    combo["values"] = opciones
    combo.current(0)

def mostrar_gastos_por_contrato(tree, tree_total, id_contrato=None, mes=None, año=None):
    datos = obtener_gastos_por_contrato(id_contrato, mes, año)

    for item in tree.get_children():
        tree.delete(item)
    for item in tree_total.get_children():
        tree_total.delete(item)

    resumen = {}
    totales_por_tipo = {"Maquinaria": 0, "Equipo y/o Htas": 0, "Servicios": 0, "Otros": 0, "Subtotal": 0, "IVA":0}

    for id_solicitud, tipo, importe, fecha, concepto, proveedor, subtotal, iva, total in datos:
        if id_solicitud not in resumen:
            resumen[id_solicitud] = {
                "Maquinaria": 0,
                "Equipo y/o Htas": 0,
                "Servicios": 0,
                "Otros": 0,
                "Fecha": fecha,
                "Concepto": concepto,
                "Proveedor": proveedor,
                "Subtotal": subtotal,
                "IVA": iva,
                "Total": total
            }
        resumen[id_solicitud][tipo] += importe
        totales_por_tipo[tipo] += importe

    for id_sol, valores in resumen.items():
        tree.insert("", "end", values=(
            id_sol,
            valores["Maquinaria"],
            valores["Equipo y/o Htas"],
            valores["Servicios"],
            valores["Otros"],
            valores["Fecha"],
            valores["Concepto"],
            valores["Proveedor"],
            valores["Subtotal"],
            valores["IVA"],
            valores["Total"]
        ))

    total_general = sum(totales_por_tipo.values())


    total_general = (
        totales_por_tipo["Maquinaria"] +
        totales_por_tipo["Equipo y/o Htas"] +
        totales_por_tipo["Servicios"] +
        totales_por_tipo["Otros"]
    )

    tree_total.insert("", "end", values=(
    totales_por_tipo["Maquinaria"],
    totales_por_tipo["Equipo y/o Htas"],
    totales_por_tipo["Servicios"],
    totales_por_tipo["Otros"],
    total_general
    ))

def exportar_reporte_excel(tree, tree_total, nombre_contrato="Todos los contratos", mes=None, año=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Gastos"

    # Estilos reutilizables
    encabezado_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    encabezado_font = Font(color="FFFFFF", bold=True)
    celda_centrada = Alignment(horizontal="center", vertical="center")
    borde = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    fila_actual = 1

    # ✅ Título del reporte si hay mes y año
    if mes and año and mes != "Todos los meses" and año != "Todos los años":
        ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=11)
        celda_titulo = ws.cell(row=fila_actual, column=1, value=f"Reporte de costos del periodo {mes} de {año}")
        celda_titulo.font = Font(bold=True)
        celda_titulo.alignment = celda_centrada
        fila_actual += 2  # Deja una fila en blanco

    # Encabezados
    encabezados = [
        "Solicitud", "Maquinaria", "Equipo y/o Htas", "Servicios", "Otros",
        "Fecha de Solicitud", "Concepto", "Proveedor", "Subtotal", "IVA", "Total", "Estado"
    ]
    for col_num, encabezado in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila_actual, column=col_num)
        celda.value = encabezado
        celda.fill = encabezado_fill
        celda.font = encabezado_font
        celda.alignment = celda_centrada
        celda.border = borde

    # Insertar datos del Treeview
    for row_idx, item in enumerate(tree.get_children(), start=fila_actual + 1):
        valores = tree.item(item, "values")
        if valores:
            id_solicitud = valores[0]

            # Insertar los valores originales
            for col_idx, valor in enumerate(valores, start=1):
                try:
                    valor_float = float(valor)
                    celda = ws.cell(row=row_idx, column=col_idx, value=valor_float)
                    celda.number_format = '"$"#,##0.00'
                except (ValueError, TypeError):
                    celda = ws.cell(row=row_idx, column=col_idx, value=valor)
                celda.alignment = celda_centrada
                celda.border = borde

            # Consultar el estado desde la base de datos
            try:
                conexion = conectar_bd()
                cursor = conexion.cursor()
                cursor.execute("""
                    SELECT estado
                    FROM solicitudespago
                    WHERE id_solicitud = %s
                """, (id_solicitud,))
                resultado = cursor.fetchone()
                estado = resultado[0] if resultado else "Desconocido"
                cursor.close()
                conexion.close()
            except Exception as e:
                estado = "Error"

            # Insertar el estado en la última columna
            celda_estado = ws.cell(row=row_idx, column=len(encabezados), value=estado)
            celda_estado.alignment = celda_centrada
            celda_estado.border = borde

    # Leer totales desde tree_total
    for item in tree_total.get_children():
        totales = tree_total.item(item, "values")
        if totales and len(totales) >= 5:
            try:
                maquinaria = float(totales[0])
                equipo = float(totales[1])
                servicios = float(totales[2])
                otros = float(totales[3])
                total_general = float(totales[4])
            except (ValueError, TypeError):
                continue

            fila_total = ws.max_row + 2  # Dos filas después del contenido

            # Etiqueta de totales
            ws.cell(row=fila_total, column=1, value=f"Totales para: {nombre_contrato}").font = Font(bold=True)

        # Clasificación por moneda: MXN y USD
    totales_por_moneda = {"MXN": [0, 0, 0, 0], "USD": [0, 0, 0, 0]}  # maquinaria, equipo, servicios, otros

    for item in tree.get_children():
        valores = tree.item(item, "values")
        if valores:
            id_solicitud = valores[0]
            try:
                maquinaria = float(valores[1])
                equipo = float(valores[2])
                servicios = float(valores[3])
                otros = float(valores[4])

                # Obtener la moneda de la solicitud
                conexion = conectar_bd()
                cursor = conexion.cursor()
                cursor.execute("""
                    SELECT ac.moneda
                    FROM solicitudespago sp
                    JOIN autorizacionescompra ac ON sp.id_autorizacion = ac.id_autorizacion
                    WHERE sp.id_solicitud = %s
                """, (id_solicitud,))
                resultado = cursor.fetchone()
                cursor.close()
                conexion.close()

                moneda = resultado[0] if resultado else "MXN"
                if moneda in totales_por_moneda:
                    totales_por_moneda[moneda][0] += maquinaria
                    totales_por_moneda[moneda][1] += equipo
                    totales_por_moneda[moneda][2] += servicios
                    totales_por_moneda[moneda][3] += otros

            except Exception as e:
                continue

    # Escribir resumen por moneda al final del Excel
    for moneda, valores in totales_por_moneda.items():
        total_general = sum(valores)
        fila_total = ws.max_row + 2

        # Título por moneda
        ws.cell(row=fila_total, column=1, value=f"Totales en {moneda}").font = Font(bold=True)

        # Encabezado "Total del contrato"
        celda_encabezado = ws.cell(row=fila_total, column=6, value="Total del contrato")
        celda_encabezado.font = encabezado_font
        celda_encabezado.fill = encabezado_fill
        celda_encabezado.alignment = celda_centrada
        celda_encabezado.border = borde

        # Fila de totales
        fila = [f"Moneda: {moneda}", *valores, total_general]
        for col_idx, valor in enumerate(fila, start=1):
            celda = ws.cell(row=fila_total + 1, column=col_idx, value=valor)
            celda.alignment = celda_centrada
            celda.border = borde
            if isinstance(valor, (int, float)):
                celda.number_format = '"$"#,##0.00'
    
    # Ajustar ancho automático de columnas (evitando MergedCell error)
    for i, col in enumerate(ws.iter_cols(min_row=1, max_col=ws.max_column), start=1):
        max_length = 0
        for cell in col:
            if cell.value and isinstance(cell, Cell):
                max_length = max(max_length, len(str(cell.value)))
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = max_length + 2

    # Diálogo para guardar archivo
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Archivos de Excel", "*.xlsx")],
        title="Guardar Reporte de Gastos"
    )

    if file_path:
        try:
            wb.save(file_path)
            messagebox.showinfo("Éxito", f"Reporte guardado correctamente en:\n{file_path}")
            os.startfile(file_path)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el archivo:\n{e}")
    else:
        messagebox.showwarning("Cancelado", "Guardado cancelado por el usuario.")

def hex_a_rgb(hex_color):
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

def crear_degradado_vertical(canvas, ancho, alto, color_inicio, color_fin):
    canvas.delete("degradado")  # Limpiar el canvas antes de dibujar un nuevo degradado

    rgb_inicio = hex_a_rgb(color_inicio)
    rgb_fin = hex_a_rgb(color_fin)

    pasos = alto // 2
    for i in range(pasos):
        y = alto - i - 1
        r = int(rgb_inicio[0] + (rgb_fin[0] - rgb_inicio[0]) * i / pasos)
        g = int(rgb_inicio[1] + (rgb_fin[1] - rgb_inicio[1]) * i / pasos)
        b = int(rgb_inicio[2] + (rgb_fin[2] - rgb_inicio[2]) * i / pasos)
        color = f"#{r:02x}{g:02x}{b:02x}"
        canvas.create_line(0, y, ancho, y, fill=color, tags="degradado")

    canvas.create_rectangle(0, 0, ancho, alto // 2, fill=color_fin, outline="", tags="degradado")

def costos_contrato():
    ventana = tk.Toplevel()
    ventana.title("Costos por Contrato")
    centrar_ventana(ventana, 1200, 600)
    
    canvas = tk.Canvas(ventana)
    canvas.pack(fill="both", expand=True)

    def actualizar_degradado(event):
        # Obtener las dimensiones del canvas
        ancho = canvas.winfo_width()
        alto = canvas.winfo_height()
        crear_degradado_vertical(canvas, ancho, alto, "#8B0000", "#FFFFFF")

    # Actualizar el fondo degradado al cambiar el tamaño de la ventana
    canvas.bind("<Configure>", actualizar_degradado)

    # Inicializar el degradado en el tamaño actual de la ventana
    ventana.after(100, lambda: actualizar_degradado(None))

    RUTA_LOGO = ruta_relativa("Plantillas/LogoATM.png")
    RUTA_LOGO2 = ruta_relativa("Plantillas/ISO-9001.jpeg")
    RUTA_LOGO3 = ruta_relativa("Plantillas/ISO-14001.jpeg")
    RUTA_LOGO4 = ruta_relativa("Plantillas/ISO-45001.jpeg")
    try:
        # LOGO ATM
        imagen = Image.open(RUTA_LOGO)
        imagen = imagen.resize((120, 160), Image.Resampling.LANCZOS)
        logo_img = ImageTk.PhotoImage(imagen)
        label_logo = tk.Label(canvas, image=logo_img, borderwidth=0)
        label_logo.image = logo_img
        label_logo.place(relx=0.07, rely=0.01)

        # ISO 9001
        iso1 = Image.open(RUTA_LOGO2).resize((70, 70), Image.Resampling.LANCZOS)
        iso1_img = ImageTk.PhotoImage(iso1)
        label_iso1 = tk.Label(canvas, image=iso1_img, borderwidth=0, bg="#ffffff")
        label_iso1.image = iso1_img
        label_iso1.place(relx=0.90, rely=0.01, anchor="ne")  # Esquina inferior derecha

        # ISO 14001
        iso2 = Image.open(RUTA_LOGO3).resize((70, 70), Image.Resampling.LANCZOS)
        iso2_img = ImageTk.PhotoImage(iso2)
        label_iso2 = tk.Label(canvas, image=iso2_img, borderwidth=0, bg="#ffffff")
        label_iso2.image = iso2_img
        label_iso2.place(relx=0.95, rely=0.15, anchor="ne")  # Al lado izquierdo del ISO 9001

        # ISO 45001
        iso3 = Image.open(RUTA_LOGO4).resize((70, 70), Image.Resampling.LANCZOS)
        iso3_img = ImageTk.PhotoImage(iso3)
        label_iso3 = tk.Label(canvas, image=iso3_img, borderwidth=0, bg="#ffffff")
        label_iso3.image = iso3_img
        label_iso3.place(relx=0.85, rely=0.15, anchor="ne")  # Al lado izquierdo del ISO 14001

    except Exception as e:
        print(f"⚠️ No se pudo cargar el logotipo: {e}")
        label_logo = tk.Label(canvas, text="LOGO", font=("Arial", 20, "bold"))


    label_titulo = tk.Label(canvas, text="ATM | Costos por Contrato", font=("Arial", 20, "bold"), bg="white")
    label_titulo.place(relx=0.5, rely=0.02, anchor="n")

    # Label y combobox de contrato
    tk.Label(canvas, text="Contrato:", font=("Arial", 10, "bold"), bg="white").place(relx=0.23, rely=0.25)
    combo_contratos = ttk.Combobox(canvas, state="readonly", width=20)
    combo_contratos.place(relx=0.30, rely=0.25)

    # Diccionario de meses
    MESES = {
        "Todos los meses": None,
        "Enero": 1, "Febrero": 2, "Marzo": 3, "Abril": 4,
        "Mayo": 5, "Junio": 6, "Julio": 7, "Agosto": 8,
        "Septiembre": 9, "Octubre": 10, "Noviembre": 11, "Diciembre": 12
    }

    # Label y combobox de mes
    tk.Label(canvas, text="Mes:", font=("Arial", 10, "bold"), bg="white").place(relx=0.50, rely=0.2)
    combo_mes = ttk.Combobox(canvas, values=list(MESES.keys()), state="readonly", width=15)
    combo_mes.place(relx=0.55, rely=0.2, relwidth=0.1)
    combo_mes.set("Todos los meses")

    # Label y combobox de año
    from datetime import datetime
    año_actual = datetime.now().year
    años = ["Todos los años"] + [str(a) for a in range(año_actual - 5, año_actual + 2)]

    tk.Label(canvas, text="Año:", font=("Arial", 10, "bold"), bg="white").place(relx=0.50, rely=0.25)
    combo_año = ttk.Combobox(canvas, values=años, state="readonly", width=10)
    combo_año.place(relx=0.55, rely=0.25, relwidth=0.1)
    combo_año.set("Todos los años")

    # Función del botón Filtrar
    def aplicar_filtro():
        nombre_contrato = combo_contratos.get()
        contrato = None if nombre_contrato == "Todos los contratos" else nombre_contrato

        mes = MESES.get(combo_mes.get(), None)
        año = None if combo_año.get() == "Todos los años" else int(combo_año.get())

        mostrar_gastos_por_contrato(tree, tree_total, contrato, mes, año)

    # Botón para aplicar el filtro
    tk.Button(canvas, text="Filtrar", font=("Arial", 10, "bold"), command=aplicar_filtro).place(relx=0.67, rely=0.22)
    # Botón para exportar a Excel
    mes_nombre = combo_mes.get()
    mes = MESES.get(mes_nombre) if mes_nombre != "Todos los meses" else None
    año_texto = combo_año.get()
    año = int(año_texto) if año_texto != "Todos los años" else None

    tk.Button(ventana, text="Exportar a Excel", font=("Arial", 10, "bold"),
        command=lambda: exportar_reporte_excel(tree, tree_total, combo_contratos.get(), mes, año)
    ).place(relx=0.78, rely=0.90)
    tk.Button(ventana, text="Salir", command= ventana.destroy, bg="red", fg="white", font=("Arial", 10, "bold")).place(relx=0.05, rely=0.91, relwidth=0.095, relheight=0.05)
    
    # Cargar contratos en el combobox
    cargar_contratos(combo_contratos)
    
    #Estilo de la tabla
    style = ttk.Style()
    style.theme_use("alt")
    style.configure("Treeview.Heading", font=("Arial", 10, "bold"), foreground="white", background="#990000")
    style.map("Treeview.Heading", background=[("!active", "#990000"), ("active", "#990000"), ("pressed", "#990000")],
              foreground=[("!active", "white"), ("active", "white"), ("pressed", "white")])


    tree_frame = ttk.Frame(canvas)
    tree_frame.place(relx=0.025, rely=0.30, relwidth=0.95, relheight=0.5)

    # Identificadores únicos para cada tabla
    columnas_tree = ("Solicitud1", "Maquinaria1", "Equipo1", "Servicios1", "Otros1", "Fecha1", "Concepto1")
    tree = ttk.Treeview(tree_frame, columns=columnas_tree, show="headings")

    # Encabezados visibles
    encabezados_visibles = ("Solicitud", "Maquinaria", "Equipo y/o Htas", "Servicios", "Otros", "Fecha de Solicitud", "Concepto", "Proveedor","Subtotal", "IVA")
    for ident, visible in zip(columnas_tree, encabezados_visibles):
        tree.heading(ident, text=visible)
        if ident in ("Solicitud1","Maquinaria1", "Equipo1", "Servicios1", "Otros1"):
            tree.column(ident, width=60, anchor="center")
        elif ident == "Fecha1":
            tree.column(ident, width=80, anchor="center")
        else:
            tree.column(ident, width=305, anchor="w")
    tree.pack(fill="both", expand=True)

    # Encabezados específicos solo para el total
    columnas_total = ("Maquinaria", "Equipo y/o Htas", "Servicios", "Otros", "TOTAL")

    tree_total = ttk.Treeview(canvas, columns=columnas_total, show="headings", height=1)
    for col in columnas_total:
        tree_total.heading(col, text=col)
        tree_total.column(col, anchor="center", width=125)

    tree_total.place(relx=0.05, rely=0.83)

    cargar_contratos(combo_contratos)

    def on_seleccion(event):
        nombre_contrato = combo_contratos.get()
        contrato = None if nombre_contrato == "Todos los contratos" else nombre_contrato
        mostrar_gastos_por_contrato(tree, tree_total, contrato)

    combo_contratos.bind("<<ComboboxSelected>>", on_seleccion)
    mostrar_gastos_por_contrato(tree, tree_total, None)
    ventana.mainloop()
