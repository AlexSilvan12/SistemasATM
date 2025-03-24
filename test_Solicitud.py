import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook
import mysql.connector
from database import conectar_bd
import os

# Ruta a la plantilla de Excel
TEMPLATE_PATH = "Plantillas\Solicitud_Pago.xlsx"

# Función para cargar autorizaciones en el Treeview
def cargar_autorizaciones(tree):
    tree.delete(*tree.get_children())  # Limpiar tabla

    try:
        conexion = conectar_bd()
        cursor = conexion.cursor()
        cursor.execute("SELECT id_autorizacion, fecha_solicitud, monto, proyecto_contrato FROM autorizacionescompra")
        for row in cursor.fetchall():
            tree.insert("", tk.END, values=row)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo cargar autorizaciones: {e}")
    finally:
        if cursor: cursor.close()
        if conexion: conexion.close()

# Función principal para generar el Excel desde la selección del Treeview
def generar_excel_desde_seleccion(tree, entry_consecutivo):

    id_solicitud = entry_consecutivo.get().strip()

    if not id_solicitud:
        messagebox.showwarning("Consecutivo vacío", "Debes ingresar un consecutivo para la solicitud.")
        return

    selected = tree.selection()
    if not selected:
        messagebox.showwarning("Atención", "Seleccione una autorización.")
        return

    id_autorizacion = tree.item(selected[0], "values")[0]

    try:
        conexion = conectar_bd()
        cursor = conexion.cursor()

        # Autorización
        cursor.execute("""
            SELECT fecha_solicitud, monto, proyecto_contrato, instruccion, id_proveedor 
            FROM autorizacionescompra 
            WHERE id_autorizacion = %s
        """, (id_autorizacion,))
        autorizacion = cursor.fetchone()
        if not autorizacion:
            messagebox.showerror("Error", "No se encontró la autorización.")
            return

        fecha_solicitud, monto, proyecto_contrato, instruccion, id_proveedor = autorizacion

        # Proveedor
        cursor.execute("""
            SELECT nombre, rfc, email, clave_bancaria, cuenta_bancaria, banco
            FROM proveedores
            WHERE id_proveedor = %s
        """, (id_proveedor,))
        proveedor = cursor.fetchone()
        if not proveedor:
            messagebox.showerror("Error", "No se encontró el proveedor.")
            return

        # Artículos
        cursor.execute("""
            SELECT cantidad, unidad, articulo, observaciones
            FROM articulosautorizacion
            WHERE id_autorizacion = %s
        """, (id_autorizacion,))
        articulos = cursor.fetchall()

        cursor.close()
        conexion.close()

        # Generar el archivo Excel
        generar_excel(id_solicitud, fecha_solicitud, monto, proyecto_contrato, instruccion, *proveedor, articulos)

    except mysql.connector.Error as e:
        messagebox.showerror("Error", f"No se pudo generar la solicitud: {e}")

# Función que llena la plantilla Excel con los datos
def generar_excel(id_solicitud, fecha_solicitud, monto, proyecto_contrato, instruccion,
                  nombre, rfc, email, clave_bancaria, cuenta_bancaria, banco, articulos):

    try:
        wb = load_workbook(TEMPLATE_PATH)
        sheet = wb.active

        # Descombinar celdas antes de escribir
        def escribir(fila, columna, valor):
            for r in sheet.merged_cells.ranges:
                if sheet.cell(row=fila, column=columna).coordinate in r:
                    sheet.unmerge_cells(str(r))
                    break
            cell = sheet.cell(row=fila, column=columna)
            cell.value = valor

        # 🧾 Llenar celdas
        escribir(6, 10, id_solicitud)           # J6
        escribir(9, 3, fecha_solicitud)         # C9
        escribir(9, 8, monto)                   # H9
        escribir(33, 8, proyecto_contrato)      # H33
        escribir(12, 3, nombre)                 # C12
        escribir(15, 7, rfc)                    # G15
        escribir(18, 7, email)                  # G18
        escribir(18, 3, clave_bancaria)         # C18
        escribir(22, 3, cuenta_bancaria)        # C22
        escribir(22, 7, banco)                  # G22
        escribir(15, 3, instruccion)            # C15

        # 🧾 Llenar artículos (fila 26 en adelante)
        fila_inicio = 25
        for i, (cantidad, unidad, articulo, observaciones) in enumerate(articulos):
            escribir(fila_inicio + i, 4, cantidad)       # B
            escribir(fila_inicio + i, 5, unidad)         # C
            escribir(fila_inicio + i, 6, articulo)       # D
            escribir(fila_inicio + i, 8, observaciones)  # G

        # Guardar y abrir el archivo
        nombre_archivo = f"C:/Sistema_SPagos/Solicitudes/solicitud_{id_solicitud}.xlsx"
        wb.save(nombre_archivo)

        messagebox.showinfo("✅ Éxito", f"Archivo generado: {nombre_archivo}")
        os.startfile(nombre_archivo)

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo generar el Excel: {e}")

# Interfaz gráfica
def interfaz():
    ventana = tk.Tk()
    ventana.title("Solicitudes de Pago")
    ventana.geometry("950x500")

    frame_filtro = tk.Frame(ventana)
    frame_filtro.pack(pady=10)

    tk.Label(frame_filtro, text="Buscar:").pack(side="left", padx=5)
    entry_filtro = tk.Entry(frame_filtro)
    entry_filtro.pack(side="left", padx=5)

    # NUEVO: Campo para ingresar el consecutivo
    frame_consecutivo = tk.Frame(ventana)
    frame_consecutivo.pack(pady=5)

    tk.Label(frame_consecutivo, text="Consecutivo de Solicitud:").pack(side="left", padx=5)
    entry_consecutivo = tk.Entry(frame_consecutivo)
    entry_consecutivo.pack(side="left", padx=5)


    tree = ttk.Treeview(ventana, columns=("ID", "Fecha", "Monto", "Proyecto"), show="headings")
    for col in ("ID", "Fecha", "Monto", "Proyecto"):
        tree.heading(col, text=col)
    tree.pack(fill="both", expand=True, padx=10, pady=10)

    btn_generar = tk.Button(ventana, text="Generar Excel", command=lambda: generar_excel_desde_seleccion(tree, entry_consecutivo))
    btn_generar.pack(pady=10)

    cargar_autorizaciones(tree)
    ventana.mainloop()

# Ejecutar la app
if __name__ == "__main__":
    interfaz()
