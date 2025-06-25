import tkinter as tk
from tkinter import PhotoImage, ttk, messagebox
from database import conectar_bd
from utils import ruta_relativa, centrar_ventana, salir
from PIL import Image, ImageTk
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl import load_workbook
from login import usuario_actual
from utils import convertir_excel_a_pdf
import mysql.connector
import os

# Funciones para cargar los datos en los treeview (puedes moverlas a tu módulo si gustas)
def cargar_autorizaciones_pendientes(tree):
    conexion = conectar_bd()
    cursor = conexion.cursor()

    try:
        consulta = """
        SELECT 
            ac.id_autorizacion, 
            ac.tipo_solicitud, 
            ac.solicitante, 
            ac.monto, 
            ac.fecha_requerida,
            GROUP_CONCAT(aa.articulo SEPARATOR ', ') AS descripcion,
            GROUP_CONCAT(aa.observaciones SEPARATOR ', ') AS observaciones
        FROM 
            autorizacionescompra ac
        LEFT JOIN 
            articulosautorizacion aa ON ac.id_autorizacion = aa.id_autorizacion
        WHERE 
            ac.estado = 'Pendiente'
        GROUP BY 
            ac.id_autorizacion, ac.tipo_solicitud, ac.solicitante, ac.monto, ac.fecha_requerida
        """

        cursor.execute(consulta)
        autorizaciones = cursor.fetchall()

        # Limpiar primero
        for item in tree.get_children():
            tree.delete(item)

        # Insertar filas
        for aut in autorizaciones:
            tree.insert("", tk.END, values=aut)

    except mysql.connector.Error as e:
        messagebox.showerror("Error", f"No se pudo cargar autorizaciones: {e}")
    finally:
        cursor.close()
        conexion.close()

def cargar_solicitudes_pendientes(tree):
    tree.delete(*tree.get_children())
    conexion = conectar_bd()
    cursor = conexion.cursor()
    try:
        consulta = """
        SELECT 
            sp.id_solicitud, 
            sp.importe, 
            sp.fecha_solicitud, 
            sp.concepto,
            GROUP_CONCAT(c.contrato SEPARATOR ', ') AS contratos
        FROM 
            solicitudespago sp
        LEFT JOIN 
            solicitud_contratos sc ON sp.id_solicitud = sc.id_solicitud
        LEFT JOIN 
            contratos c ON sc.id_contrato = c.id_contrato
        WHERE 
            sp.estado = 'Pendiente'
        GROUP BY 
            sp.id_solicitud, sp.importe, sp.fecha_solicitud, sp.concepto 
        """
        cursor.execute(consulta)
        solicitudes = cursor.fetchall()

         # Limpiar primero
        for item in tree.get_children():
            tree.delete(item)

        # Insertar filas
        for aut in solicitudes:
            tree.insert("", tk.END, values=aut)   

    except mysql.connector.Error as e:
        messagebox.showerror("Error", f"No se pudo cargar solicitudes: {e}")

    finally:
        cursor.close()
        conexion.close()

# Funcion para autorizar las autorizaciones de compras
def autorizar_autorizacion_y_solicitud(tree):
    seleccion = tree.selection()
    if not seleccion:
        messagebox.showwarning("Selecciona una autorización", "Por favor selecciona una autorización para autorizar.")
        return

    id_autorizacion = tree.item(seleccion, "values")[0]

    if not messagebox.askyesno("Confirmar", f"¿Deseas autorizar la compra y su solicitud relacionada ({id_autorizacion})?"):
        return

    conexion = conectar_bd()
    cursor = conexion.cursor()

    try:
        # Obtener el ID de la solicitud relacionada con la autorización
        cursor.execute("""
            SELECT sp.id_solicitud 
            FROM solicitudespago sp
            JOIN autorizacionescompra ac ON ac.id_autorizacion = sp.id_autorizacion
            WHERE ac.id_autorizacion = %s
        """, (id_autorizacion,))
        resultado = cursor.fetchone()

        if not resultado:
            raise Exception("No se encontró una solicitud relacionada con esta autorización.")

        id_solicitud = resultado[0]

        # Autorizar ambas en la base de datos
        cursor.execute("UPDATE autorizacionescompra SET estado = 'Autorizado' WHERE id_autorizacion = %s", (id_autorizacion,))
        cursor.execute("UPDATE solicitudespago SET estado = 'Autorizado' WHERE id_solicitud = %s", (id_solicitud,))
        conexion.commit()

        # Rutas de archivos
        ruta_excel_aut = os.path.join(os.environ['USERPROFILE'], 'OneDrive - ATI', 'Documentos de AppGestor', 'Autorizaciones de Compra', f'Autorizacion_{id_autorizacion}.xlsx')
        ruta_pdf_aut = ruta_excel_aut.replace(".xlsx", ".pdf")

        ruta_excel_sol = os.path.join(os.environ['USERPROFILE'], 'OneDrive - ATI', 'Documentos de AppGestor', 'Solicitudes de Pago', f'Solicitud de Pago_{id_solicitud}.xlsx')
        ruta_pdf_sol = ruta_excel_sol.replace(".xlsx", ".pdf")

        ruta_firma = ruta_relativa(usuario_actual["firma"])
        firma_img = ExcelImage(ruta_firma)
        firma_img.width = 150
        firma_img.height = 50

        # Agregar firma a archivo de autorización
        if os.path.exists(ruta_excel_aut):
            wb_aut = load_workbook(ruta_excel_aut)
            sheet_aut = wb_aut.active
            sheet_aut.add_image(firma_img, "G37")
            wb_aut.save(ruta_excel_aut)
            convertir_excel_a_pdf(ruta_excel_aut, ruta_pdf_aut)

        # Agregar firma a archivo de solicitud
        if os.path.exists(ruta_excel_sol):
            wb_sol = load_workbook(ruta_excel_sol)
            sheet_sol = wb_sol.active
            sheet_sol.add_image(firma_img, "I37")
            wb_sol.save(ruta_excel_sol)
            convertir_excel_a_pdf(ruta_excel_sol, ruta_pdf_sol)

        messagebox.showinfo("✅ Autorizado", f"La autorización {id_autorizacion} y su solicitud {id_solicitud} fueron autorizadas correctamente.")
        cargar_autorizaciones_pendientes(tree)

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo autorizar: {e}")
        return

    finally:
        cursor.close()
        conexion.close()
def mostrar_detalles_autorizacion(id_autorizacion):
    conexion = conectar_bd()
    cursor = conexion.cursor()

    try:
        cursor.execute("""
            SELECT articulo, observaciones
            FROM articulosautorizacion
            WHERE id_autorizacion = %s
        """, (id_autorizacion,))
        articulos = cursor.fetchall()

        if not articulos:
            messagebox.showinfo("Sin datos", "No hay artículos registrados para esta autorización.")
            return

        # Crear ventana emergente
        detalle_ventana = tk.Toplevel()
        detalle_ventana.title(f"Detalles de Autorización {id_autorizacion}")
        detalle_ventana.geometry("600x400")

        # Text widget para mostrar los datos
        texto = tk.Text(detalle_ventana, wrap="word", font=("Arial", 11))
        texto.pack(fill="both", expand=True, padx=10, pady=10)

        for art, obs in articulos:
            texto.insert("end", f"🛒 Artículo: {art}\n📝 Observaciones: {obs}\n\n")

        texto.config(state="disabled")  # Solo lectura

    except mysql.connector.Error as e:
        messagebox.showerror("Error", f"No se pudieron obtener los detalles:\n{e}")
    finally:
        cursor.close()
        conexion.close()

#Interfaz de Usuario
def hex_a_rgb(hex_color):
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2 ,4))

def crear_degradado_vertical(canvas, ancho, alto, color_inicio, color_fin):
    canvas.delete("degradado")

    rgb_inicio = hex_a_rgb(color_inicio)
    rgb_fin = hex_a_rgb(color_fin)

    pasos = alto // 2
    for i in range(pasos):
        y = alto - i - 1
        r = int(rgb_inicio[0] + (rgb_fin[0] - rgb_inicio[0]) * i / pasos)
        g = int(rgb_inicio[1] + (rgb_fin[1] - rgb_inicio[1]) * i / pasos)
        b = int(rgb_inicio[2] + (rgb_fin[2] - rgb_inicio[2]) * i / pasos)
        color = f"#{r:02x}{g:02x}{b:02x}"
        canvas.create_line(0, y, ancho, y, fill=color, tags="degradado")

    canvas.create_rectangle(0, 0, ancho, alto // 2, fill=color_fin, outline="", tags="degradado")


def Autorizacion_Pagos_Compras(rol, volver_menu_callback):
    ventana = tk.Tk()
    ventana.title("Gestión de Solicitudes y Autorizaciones")
    centrar_ventana(ventana, 1100, 600)

    canvas = tk.Canvas(ventana)
    canvas.pack(fill="both", expand=True)

    def actualizar_degradado(event):
        ancho = canvas.winfo_width()
        alto = canvas.winfo_height()
        crear_degradado_vertical(canvas, ancho, alto, "#8B0000", "#FFFFFF")

    canvas.bind("<Configure>", actualizar_degradado)
    ventana.after(100, lambda: actualizar_degradado(None))

    # Cargar logos
    RUTA_LOGO = ruta_relativa("Plantillas/LogoATM.png")
    RUTA_LOGO2 = ruta_relativa("Plantillas/ISO-9001.jpeg")
    RUTA_LOGO3 = ruta_relativa("Plantillas/ISO-14001.jpeg")
    RUTA_LOGO4 = ruta_relativa("Plantillas/ISO-45001.jpeg")
    try:
        # LOGO ATM
        imagen = Image.open(RUTA_LOGO)
        imagen = imagen.resize((150, 160), Image.Resampling.LANCZOS)
        logo_img = ImageTk.PhotoImage(imagen)
        label_logo = tk.Label(canvas, image=logo_img, borderwidth=0)
        label_logo.image = logo_img
        label_logo.place(relx=0.07, rely=0.01)

        # ISO 9001
        iso1 = Image.open(RUTA_LOGO2).resize((70, 70), Image.Resampling.LANCZOS)
        iso1_img = ImageTk.PhotoImage(iso1)
        label_iso1 = tk.Label(canvas, image=iso1_img, borderwidth=0, bg="#ffffff")
        label_iso1.image = iso1_img
        label_iso1.place(relx=0.90, rely=0.01, anchor="ne")  # Esquina inferior derecha

        # ISO 14001
        iso2 = Image.open(RUTA_LOGO3).resize((70, 70), Image.Resampling.LANCZOS)
        iso2_img = ImageTk.PhotoImage(iso2)
        label_iso2 = tk.Label(canvas, image=iso2_img, borderwidth=0, bg="#ffffff")
        label_iso2.image = iso2_img
        label_iso2.place(relx=0.95, rely=0.15, anchor="ne")  # Al lado izquierdo del ISO 9001

        # ISO 45001
        iso3 = Image.open(RUTA_LOGO4).resize((70, 70), Image.Resampling.LANCZOS)
        iso3_img = ImageTk.PhotoImage(iso3)
        label_iso3 = tk.Label(canvas, image=iso3_img, borderwidth=0, bg="#ffffff")
        label_iso3.image = iso3_img
        label_iso3.place(relx=0.85, rely=0.15, anchor="ne")  # Al lado izquierdo del ISO 14001

    except Exception as e:
        print(f"⚠️ No se pudo cargar el logotipo: {e}")
        label_logo = tk.Label(canvas, text="LOGO", font=("Arial", 20, "bold"))
    
    label_titulo = tk.Label(canvas, text="ATM | Autorización de Compras y Pagos",
                            font=("Arial", 20, "bold"), fg="black", bg="white")
    label_titulo.place(relx=0.27, rely=0.10)

    # --- Treeview único ---
    frame_autorizaciones = tk.Frame(canvas)
    frame_autorizaciones.place(relx=0.05, rely=0.3, relwidth=0.9, relheight=0.55)

    style = ttk.Style()
    style.theme_use("alt")
    style.configure("Treeview.Heading", font=("Arial", 10, "bold"), foreground="white", background="#990000")

    tree_aut = ttk.Treeview(frame_autorizaciones, columns=("ID", "Tipo", "Solicitante", "Monto", "Fecha Requerida", "Descripcion", "Observaciones"), show="headings")
    for col in tree_aut["columns"]:
        tree_aut.heading(col, text=col)
        ancho = 400 if col in ("Descripcion", "Observaciones") else 185 if col == "Solicitante" else 100
        tree_aut.column(col, width=ancho, anchor="w" if col in ("Descripcion", "Observaciones", "Solicitante") else "center")
    tree_aut.pack(fill="both", expand=True, padx=10, pady=10)

    scrollbar_x = ttk.Scrollbar(frame_autorizaciones, orient="horizontal", command=tree_aut.xview)
    scrollbar_x.pack(side="bottom", fill="x")
    tree_aut.configure(xscrollcommand=scrollbar_x.set)

    def on_autorizacion_select(event):
        item = tree_aut.focus()
        if item:
            id_autorizacion = tree_aut.item(item, "values")[0]
            mostrar_detalles_autorizacion(id_autorizacion)

    tree_aut.bind("<Double-1>", on_autorizacion_select)

    def ventana_autorizados():
        ventana = tk.Toplevel()
        ventana.title("Autorizaciones y Solicitudes Autorizadas")
        centrar_ventana(ventana, 1100, 600)

        canvas = tk.Canvas(ventana)
        canvas.pack(fill="both", expand=True)

        def actualizar_degradado(event):
            # Obtener las dimensiones del canvas
            ancho = canvas.winfo_width()
            alto = canvas.winfo_height()
            crear_degradado_vertical(canvas, ancho, alto, "#8B0000", "#FFFFFF")

        # Actualizar el fondo degradado al cambiar el tamaño de la ventana
        canvas.bind("<Configure>", actualizar_degradado)

        # Inicializar el degradado en el tamaño actual de la ventana
        ventana.after(100, lambda: actualizar_degradado(None))
        RUTA_LOGO = ruta_relativa("Plantillas/LogoATM.png")
        RUTA_LOGO2 = ruta_relativa("Plantillas/ISO-9001.jpeg")
        RUTA_LOGO3 = ruta_relativa("Plantillas/ISO-14001.jpeg")
        RUTA_LOGO4 = ruta_relativa("Plantillas/ISO-45001.jpeg")
        try:
            # LOGO ATM
            imagen = Image.open(RUTA_LOGO)
            imagen = imagen.resize((150, 160), Image.Resampling.LANCZOS)
            logo_img = ImageTk.PhotoImage(imagen)
            label_logo = tk.Label(canvas, image=logo_img, borderwidth=0)
            label_logo.image = logo_img
            label_logo.place(relx=0.07, rely=0.01)

            # ISO 9001
            iso1 = Image.open(RUTA_LOGO2).resize((70, 70), Image.Resampling.LANCZOS)
            iso1_img = ImageTk.PhotoImage(iso1)
            label_iso1 = tk.Label(canvas, image=iso1_img, borderwidth=0, bg="#ffffff")
            label_iso1.image = iso1_img
            label_iso1.place(relx=0.90, rely=0.015, anchor="ne")  # Esquina inferior derecha

            # ISO 14001
            iso2 = Image.open(RUTA_LOGO3).resize((70, 70), Image.Resampling.LANCZOS)
            iso2_img = ImageTk.PhotoImage(iso2)
            label_iso2 = tk.Label(canvas, image=iso2_img, borderwidth=0, bg="#ffffff")
            label_iso2.image = iso2_img
            label_iso2.place(relx=0.95, rely=0.17, anchor="ne")  # Al lado izquierdo del ISO 9001

            # ISO 45001
            iso3 = Image.open(RUTA_LOGO4).resize((70, 70), Image.Resampling.LANCZOS)
            iso3_img = ImageTk.PhotoImage(iso3)
            label_iso3 = tk.Label(canvas, image=iso3_img, borderwidth=0, bg="#ffffff")
            label_iso3.image = iso3_img
            label_iso3.place(relx=0.85, rely=0.17, anchor="ne")  # Al lado izquierdo del ISO 14001

        except Exception as e:
            print(f"⚠️ No se pudo cargar el logotipo: {e}")
            label_logo = tk.Label(canvas, text="LOGO", font=("Arial", 20, "bold"))

        label_titulo = tk.Label(canvas, text="ATM | Compras y Pagos Autorizados",
                            font=("Arial", 20, "bold"), fg="black", bg="white")
        label_titulo.place(relx=0.27, rely=0.10)


        notebook = ttk.Notebook(ventana)
        notebook.place(relx=0.05, rely=0.3, relwidth=0.9, relheight=0.6)


        # --- Pestaña Autorizaciones Autorizadas ---
        frame_autorizadas = tk.Frame(ventana)
        notebook.add(frame_autorizadas, text="Autorizaciones de Compra Autorizadas")

        tree_aut_autorizadas = ttk.Treeview(
            frame_autorizadas,
            columns=("ID", "Tipo", "Solicitante", "Monto", "Fecha Requerida", "Descripcion"),
            show="headings"
        )

        # Definir columnas
        for col in tree_aut_autorizadas["columns"]:
            tree_aut_autorizadas.heading(col, text=col)
            if col == "Descripcion":
                tree_aut_autorizadas.column(col, width=400, anchor="w")
            elif col == "Solicitante":
                tree_aut_autorizadas.column(col,width=185 ,anchor="w")
            else:
                tree_aut_autorizadas.column(col, width=100, anchor="center")

        tree_aut_autorizadas.pack(side="top", fill="both", expand=True)

        # ✅ Scrollbar horizontal
        scrollbar_x2 = ttk.Scrollbar(frame_autorizadas, orient="horizontal", command=tree_aut_autorizadas.xview)
        scrollbar_x2.pack(side="bottom", fill="x")

        tree_aut_autorizadas.configure(xscrollcommand=scrollbar_x2.set)

        # Cargar datos
        cargar_autorizaciones_autorizadas(tree_aut_autorizadas)


        # --- Pestaña Solicitudes Autorizadas ---
        frame_solicitudes = tk.Frame(notebook, bg="white")
        notebook.add(frame_solicitudes, text="Solicitudes de Pago Autorizadas")

        tree_solicitudes = ttk.Treeview(frame_solicitudes, columns=("ID", "Fecha", "Importe", "Estado", "Concepto"), show="headings")
        for col in ("ID", "Fecha", "Importe", "Estado","Concepto"):
            tree_solicitudes.heading(col, text=col)
            if col == "Concepto":
                tree_solicitudes.column(col, width=400, anchor="w")
            else:
                tree_solicitudes.column(col, width=100, anchor="center")
        tree_solicitudes.pack(fill="both", expand=True, padx=10, pady=10)

        # ✅ Scrollbar horizontal
        scrollbar_x2 = ttk.Scrollbar(frame_solicitudes, orient="horizontal", command=tree_solicitudes.xview)
        scrollbar_x2.pack(side="bottom", fill="x")

        tk.Button(ventana, text="Salir", command= ventana.destroy, bg="red", fg="white", font=("Arial", 10, "bold")).place(relx=0.05, rely=0.92, relwidth=0.08, relheight=0.04)


        tree_solicitudes.configure(xscrollcommand=scrollbar_x2.set)
      
        cargar_solicitudes_autorizadas(tree_solicitudes)

# --- Funciones para cargar datos ---
    def cargar_solicitudes_autorizadas(tree):
        conexion = conectar_bd()
        cursor = conexion.cursor()
        try:
            query = """
                SELECT id_solicitud, fecha_solicitud, importe, estado, concepto
                FROM solicitudespago
                WHERE estado = 'Autorizado' OR estado = 'Pagado'
            """
            cursor.execute(query)
            resultados = cursor.fetchall()
            for row in resultados:
                tree.insert("", tk.END, values=row)
        except Exception as e:
            print(f"❌ Error al cargar solicitudes autorizadas: {e}")
        finally:
            cursor.close()
            conexion.close()

    def cargar_autorizaciones_autorizadas(tree):
        conexion = conectar_bd()
        cursor = conexion.cursor()

        consulta = """
            SELECT ac.id_autorizacion, ac.tipo_solicitud, ac.solicitante, ac.monto, ac.fecha_requerida, 
                GROUP_CONCAT(aa.articulo SEPARATOR ',')
            FROM autorizacionescompra ac
            LEFT JOIN articulosautorizacion aa ON ac.id_autorizacion = aa.id_autorizacion
            WHERE ac.estado = 'Autorizado'
            GROUP BY ac.id_autorizacion
        """

        cursor.execute(consulta)
        registros = cursor.fetchall()

        for row in registros:
            tree.insert("", "end", values=row)

        conexion.close()

    # Botones
    tk.Button(canvas, text="Autorizar Compras y Pago", font=("Arial", 10, "bold"),
              command=lambda: autorizar_autorizacion_y_solicitud(tree_aut)).place(relx=0.45, rely=0.91, relwidth=0.18, relheight=0.06)

    tk.Button(canvas, text="Compras y Pagos Autorizados", font=("Arial", 10, "bold"),
              command=ventana_autorizados).place(relx=0.65, rely=0.91, relwidth=0.19, relheight=0.06)

    tk.Button(ventana, text="Salir", command=lambda: salir(volver_menu_callback, ventana),
              bg="red", fg="white", font=("Arial", 10, "bold")).place(relx=0.05, rely=0.92, relwidth=0.08, relheight=0.04)

    cargar_autorizaciones_pendientes(tree_aut)
    ventana.mainloop()
