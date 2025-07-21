import shutil
import tkinter as tk
from tkinter import filedialog
from gastos_contrato import costos_contrato
from tkinter import messagebox, ttk
from PIL import Image, ImageTk
from datetime import date
from database import conectar_bd
from utils import ruta_relativa, centrar_ventana, convertir_excel_a_pdf, salir
from login import usuario_actual
from openpyxl import load_workbook
import mysql.connector
import os
from openpyxl.styles import Alignment, Border, Side
from openpyxl.drawing.image import Image as ExcelImage

factura_seleccionada = None

# Función para conectar con las solicitudes almacenadas en la base de datos
def cargar_solicitudes(tree):
    for row in tree.get_children():
        tree.delete(row)  # Limpiar datos previos

    conexion = None
    cursor = None

    try:
        conexion = conectar_bd()
        if conexion is None:
            print("❌ No se pudo establecer la conexión")
            return

        cursor = conexion.cursor()
        cursor.execute("SELECT id_solicitud, fecha_solicitud, importe, fecha_limite_pago, estado FROM SolicitudesPago WHERE estado = 'Autorizado' OR estado = 'Pendiente'")

        for solicitud in cursor.fetchall():
            tree.insert("", "end", values=solicitud)

    except Exception as e:
        print(f"❌ Error al cargar solicitudes de pago: {e}")

    finally:
        if cursor is not None:
            cursor.close()
        if conexion is not None:
            conexion.close()

def cargar_autorizaciones(tree):
    tree.delete(*tree.get_children())  # Limpiar tabla

    try:
        conexion = conectar_bd()
        cursor = conexion.cursor()
        cursor.execute("""
    SELECT id_autorizacion, fecha_solicitud, monto, fecha_limite_pago 
    FROM autorizacionescompra
    WHERE id_autorizacion NOT IN (
        SELECT id_autorizacion FROM SolicitudesPago
    )
""")
        for row in cursor.fetchall():
            tree.insert("", tk.END, values=row)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo cargar autorizaciones: {e}")
    finally:
        if cursor: cursor.close()
        if conexion: conexion.close()

def marcar_como_pagado(tree, usuario_actual):
    if usuario_actual["rol"] != "Contador":
        messagebox.showwarning("Acceso denegado", "No tiene permisos para realizar esta acción.")
        return

    seleccion = tree.focus()
    if not seleccion:
        messagebox.showwarning("Sin selección", "Selecciona una solicitud para marcar como pagada.")
        return

    valores = tree.item(seleccion, "values")
    if not valores:
        return

    id_solicitud = valores[0]

    try:
        conexion = conectar_bd()
        cursor = conexion.cursor()

        cursor.execute("SELECT estado FROM solicitudespago WHERE id_solicitud = %s", (id_solicitud,))
        estado_actual = cursor.fetchone()
        if not estado_actual or estado_actual[0] != "Autorizado":
            messagebox.showinfo("Información", "La Solicitud de Pago aun no ha sido autorizada.")
            return

        hoy = date.today()

        cursor.execute("""
            UPDATE solicitudespago 
            SET estado = 'Pagado', fecha_pago = %s 
            WHERE id_solicitud = %s
        """, (hoy, id_solicitud))

        conexion.commit()
        messagebox.showinfo("✅ Éxito", f"La solicitud {id_solicitud} fue marcada como 'Pagado' el {hoy}.")

        # Opcional: refrescar el treeview si tienes una función como cargar_solicitudes_pendientes()
        cargar_solicitudes(tree)

    except Exception as e:
        messagebox.showerror("❌ Error", f"No se pudo actualizar el estado:\n{e}")
    finally:
        if cursor: cursor.close()
        if conexion: conexion.close()

# Función principal para generar el Excel desde la selección del Treeview
def generar_excel_desde_seleccion(tree, entry_consecutivo, entry_concepto, entry_referencia, entry_factura):
    id_solicitud = entry_consecutivo.get().strip()
    concepto = entry_concepto.get("1.0", "end").strip()
    referencia_pago = entry_referencia.get().strip()
    global factura_seleccionada

    if not factura_seleccionada:
        messagebox.showerror("Error", "Debes seleccionar un archivo PDF como factura.")
        return

    if not id_solicitud:
        messagebox.showwarning("Consecutivo vacío", "Debes ingresar un consecutivo para la solicitud.")
        return

    selected = tree.selection()
    if not selected:
        messagebox.showwarning("Atención", "Seleccione una autorización.")
        return

    id_autorizacion = tree.item(selected[0], "values")[0]

    try:
        conexion = conectar_bd()
        cursor = conexion.cursor()

        cursor.execute("SELECT COUNT(*) FROM SolicitudesPago WHERE id_solicitud = %s", (id_solicitud,))
        if cursor.fetchone()[0] > 0:
            messagebox.showwarning("Advertencia", f"Ya existe una solicitud con el ID '{id_solicitud}'. No se generó el Excel ni se guardaron datos.")
            return

        # Obtener datos de la autorización
        cursor.execute("""
            SELECT fecha_solicitud, monto, instruccion, id_proveedor, fecha_limite_pago, IVA, subtotal
            FROM autorizacionescompra 
            WHERE id_autorizacion = %s
        """, (id_autorizacion,))
        autorizacion = cursor.fetchone()
        if not autorizacion:
            messagebox.showerror("Error", "No se encontró la autorización.")
            return

        fecha_solicitud, monto, instruccion, id_proveedor, fechalimite, iva, subtotal = autorizacion

        # Obtener moneda desde la autorización
        cursor.execute("SELECT moneda FROM autorizacionescompra WHERE id_autorizacion = %s", (id_autorizacion,))
        moneda = cursor.fetchone()
        moneda = moneda[0]


        # Obtener datos del proveedor
        cursor.execute("""
            SELECT nombre, rfc, email, clave_bancaria, cuenta_bancaria, banco
            FROM proveedores
            WHERE id_proveedor = %s
        """, (id_proveedor,))
        proveedor = cursor.fetchone()
        if not proveedor:
            messagebox.showerror("Error", "No se encontró el proveedor.")
            return

        # Guardar el archivo PDF en la carpeta de facturas con el nombre correcto
        base_onedrive = os.path.join(
            os.environ['USERPROFILE'],
            'OneDrive - ATI',
            'Documentos de AppGestor'
        )
        CARPETA_FACTURAS = os.path.join(base_onedrive, 'Facturas')
        if not os.path.exists(CARPETA_FACTURAS):
            os.makedirs(CARPETA_FACTURAS)

        nombre_archivo = f"Factura_Solicitud de Pago_{id_solicitud}.pdf"
        ruta_destino = os.path.join(CARPETA_FACTURAS, nombre_archivo)

        try:
            shutil.copy(factura_seleccionada, ruta_destino)
            factura = ruta_destino
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo copiar el archivo: {e}")
            return

        # Insertar en la tabla SolicitudesPago
        query = """
            INSERT INTO SolicitudesPago (
                id_solicitud, id_autorizacion, id_proveedor, fecha_solicitud, 
                importe, instruccion, referencia_pago, concepto, 
                fecha_recibido_revision, fecha_limite_pago, num_facturas, estado, IVA, subtotal
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'Pendiente', %s, %s)
        """
        cursor.execute(query, (
            id_solicitud, id_autorizacion, id_proveedor, fecha_solicitud,
            monto, instruccion, referencia_pago, concepto, fecha_solicitud, fechalimite, factura, iva, subtotal
        ))
        conexion.commit()

        # Insertar los contratos asociados
        cursor.execute("SELECT id_contrato, importe FROM Autorizacion_Contratos WHERE id_autorizacion = %s", (id_autorizacion,))
        contratos = cursor.fetchall()
        for id_contrato, importe in contratos:
            cursor.execute("""
                INSERT INTO solicitud_contratos (id_solicitud, id_contrato, importe)
                VALUES (%s, %s, %s)
            """, (id_solicitud, id_contrato, importe))
        conexion.commit()

        # Limpiar campos y actualizar UI
        messagebox.showinfo("✅ Éxito", f"Solicitud '{id_solicitud}' guardada en la base de datos.")
        entry_consecutivo.delete(0, tk.END)
        entry_referencia.delete(0, tk.END)
        entry_concepto.delete("1.0", "end")
        entry_factura.config(state="normal")
        entry_factura.delete(0, tk.END)
        entry_factura.config(state="readonly")
        factura_seleccionada = None
        cargar_autorizaciones(tree)

    except mysql.connector.Error as err:
        messagebox.showerror("Error", f"No se pudo registrar la solicitud:\n{err}")
    finally:
        if cursor: cursor.close()
        if conexion: conexion.close()

    # Generar Excel y PDF
    generar_excel(id_solicitud, fecha_solicitud, monto, instruccion,
                  referencia_pago, fechalimite, concepto, factura,
                  *proveedor, usuario_actual["nombre"], iva, subtotal, moneda)

def seleccionar_factura(entry_factura):
    global factura_seleccionada
    if entry_factura.get().strip():
        messagebox.showwarning("Advertencia", "Ya se ha ingresado un número de factura.")
        return

    ruta_origen = filedialog.askopenfilename(title="Seleccionar Archivo de Factura", filetypes=[("Archivos PDF", "*.pdf")])
    if ruta_origen:
        factura_seleccionada = ruta_origen
        entry_factura.delete(0, tk.END)
        entry_factura.insert(0, ruta_origen)
        entry_factura.config(state='readonly')

        messagebox.showinfo("Archivo Seleccionado", f"Factura seleccionada:\n{os.path.basename(ruta_origen)}")

def eliminar_factura(entry_factura):
   global factura_seleccionada
   factura_seleccionada = None
   entry_factura.config(state='normal')  # Permitir edición
   entry_factura.delete(0, tk.END)       # Borrar contenido

   seleccionar_factura(entry_factura)
       
# Función que llena la plantilla Excel con los datos
def generar_excel(id_solicitud, fecha_solicitud, monto, instruccion,
                  referencia_pago, fechalimite, concepto, factura, nombre, rfc, email, clave_bancaria, cuenta_bancaria, banco, nombre_usuario, iva, subtotal, moneda):
  
    try:
        # Obtener los nombres de contratos asociados a la solicitud
        conexion = conectar_bd()
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT c.contrato 
            FROM solicitud_contratos sc
            JOIN contratos c ON sc.id_contrato = c.id_contrato
            WHERE sc.id_solicitud = %s
        """, (id_solicitud,))
        nombres_contratos = [fila[0] for fila in cursor.fetchall()]
        texto_contratos = ", ".join(nombres_contratos)

        # Plantilla de solicitud de pago
        plantilla_path = ruta_relativa("Plantillas/Solicitud_Pago.xlsx")
        wb = load_workbook(plantilla_path)
        sheet = wb.active

        # Función para escribir en celdas combinadas
        def escribir(fila, columna, valor, combinar=None):
            celda = sheet.cell(row=fila, column=columna)

            # Verificar si está en un rango combinado
            for r in sheet.merged_cells.ranges:
                if celda.coordinate in r:
                    sheet.unmerge_cells(str(r))
                    break

            celda.value = valor
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text= True)

            # Si se desea recombinar
            if combinar:
                sheet.merge_cells(combinar)

        # Llenar celdas con datos generales
        escribir(6, 10, id_solicitud, combinar="J6:L6")          # J6 - Consecutivo
        escribir(9, 3, fecha_solicitud, combinar="C9:E9")        # C9 - Fecha
        escribir(9, 8, subtotal, combinar="H9:I9")               # Subtotal
        escribir(9, 10, iva, combinar="J9:K9")                   # IVA
        escribir(9, 12, f"{monto} {moneda}", combinar="L9:L9")  # Total con moneda
        escribir(34, 8, texto_contratos, combinar="H34:L34")     # H34 - Proyecto

        escribir(12, 3, nombre, combinar="C12:L12")              # C12 - Nombre proveedor
        escribir(15, 7, rfc, combinar="G15:L15")                 # G15 - RFC
        escribir(18, 8, email, combinar="H18:L18")               # G18 - Email
        escribir(18, 3, clave_bancaria, combinar="C18:F18")      # C18 - Clave bancaria
        escribir(22, 3, cuenta_bancaria, combinar="C22:E22")     # C22 - Cuenta bancaria
        escribir(22, 7, banco, combinar="G22:H22")               # G22 - Banco
        escribir(29, 8, fechalimite, combinar="H29:L29")         # H29 - Limite de Pago
        escribir(15, 3, instruccion, combinar="C15:E15")         # C15 - Instrucción
        escribir(22, 10, referencia_pago, combinar="J22:L22")    # J22 - Referencia de pago
        escribir(25, 3, concepto, combinar="C25:L25")            # C25 - Concepto
        escribir(38, 3, nombre_usuario, combinar="C38:F38")      # C37 - Solicitante de Pago
        
        # Ver si la factura es un archivo PDF o texto normal
        if factura.lower().endswith(".pdf"):
            base_onedrive = os.path.join(
                os.environ['USERPROFILE'],
                    'OneDrive - ATI',
                    'Documentos de AppGestor',
                    'Facturas'
                    )
        RUTA_FACTURAS = os.path.join(base_onedrive, os.path.basename(factura))

        from openpyxl.styles import Border, Side

        if os.path.exists(RUTA_FACTURAS):
            nombre_archivo = os.path.basename(factura)
            celda_factura = sheet.cell(row=34, column=3)
            celda_factura.value = nombre_archivo
            celda_factura.hyperlink = RUTA_FACTURAS
            celda_factura.style = "Hyperlink"
            celda_factura.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Bordes a aplicar
            borde_superior_izquierdo = Border(top=Side(style="thin"), left=Side(style="thin"))
            borde_superior = Border(top=Side(style="thin"))
            borde_izquierdo = Border(left=Side(style="thin"))

            # Aplicar bordes a todas las celdas del rango C34:F34 (columnas 3 a 6)
            for col in range(3, 7):
                celda = sheet.cell(row=34, column=col)
                if col == 3:  # celda C34
                    celda.border = borde_superior_izquierdo
                elif col == 6:  # última columna F34, solo superior 
                    celda.border = borde_superior
                else:  # D34, E34
                    celda.border = borde_superior

            # Combinar las celdas después de aplicar los bordes
            sheet.merge_cells("C34:F34")
        else:
            escribir(34, 3, factura, combinar="C34:F34")

        ruta_firma = ruta_relativa(usuario_actual["firma"])

        # Insertar imagen
        firma_img = ExcelImage(ruta_firma)
        firma_img.width =160 #ajustar tamaño
        firma_img.height = 50
        sheet.add_image(firma_img, "D37")

        # Guardar Excel
        #CARPETA_SOLICITUDES = ruta_relativa("Solicitudes") #Quitar cuando se sincronicen en one drive
        CARPETA_SOLICITUDES= os.path.join(os.environ['USERPROFILE'], 'OneDrive - ATI', 'Documentos de AppGestor', 'Solicitudes de Pago')
        #Crear la carpeta si no existe
        if not os.path.exists(CARPETA_SOLICITUDES):
            os.makedirs(CARPETA_SOLICITUDES)
            
        output_path = os.path.join(CARPETA_SOLICITUDES, f"Solicitud de Pago_{id_solicitud}.xlsx")
        wb.save(output_path)

        # Convertir a PDF (la función se encarga de abrirlo)
        ruta_pdf = output_path.replace(".xlsx", ".pdf")
        convertir_excel_a_pdf(output_path, ruta_pdf)

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo generar el Excel:\n{e}")

def mSolicitud_excel(id_autorizacion):
    conexion = conectar_bd()
    cursor = conexion.cursor()

    # Buscar solicitud relacionada
    cursor.execute("SELECT * FROM SolicitudesPago WHERE id_autorizacion = %s", (id_autorizacion,))
    solicitud = cursor.fetchone()
    if not solicitud:
        messagebox.showinfo("Sin solicitud", "No hay una solicitud de pago relacionada a esta autorización.")
        return

    (
        id_solicitud, _, id_proveedor, fecha_solicitud, monto, instruccion,
        referencia_pago, concepto, _, fecha_limite, factura, _, estado, iva, subtotal
    ) = solicitud

    # Buscar datos actuales de autorización y proveedor
    cursor.execute("""
        SELECT a.fecha_solicitud, a.monto, a.instruccion, a.fecha_limite_pago,
               p.nombre, p.rfc, p.email, p.clave_bancaria, p.cuenta_bancaria, p.banco
        FROM autorizacionescompra a
        JOIN proveedores p ON a.id_proveedor = p.id_proveedor
        WHERE a.id_autorizacion = %s
    """, (id_autorizacion,))
    resultado = cursor.fetchone()
    if not resultado:
        messagebox.showerror("Error", "No se encontró la autorización o proveedor.")
        return

    nueva_fecha, nuevo_monto, nueva_instr, nueva_fecha_lim, nombre, rfc, email, clave, cuenta, banco = resultado

    # Comparar campos clave para ver si ha habido cambios
    cambios = (
        str(fecha_solicitud) != str(nueva_fecha) or
        float(monto) != float(nuevo_monto) or
        instruccion != nueva_instr or
        str(fecha_limite) != str(nueva_fecha_lim)
    )

    if not cambios:
        print("No se detectaron cambios en la solicitud de pago. No se modificará el Excel.")
        return

    # Generar nuevo Excel sobrescribiendo
    generar_excel(
        id_solicitud, nueva_fecha, nuevo_monto, nueva_instr, referencia_pago,
        nueva_fecha_lim, concepto, factura, nombre, rfc, email, clave, cuenta,
        banco, usuario_actual["nombre"], iva, subtotal
    )

    print(f"Solicitud de pago {id_solicitud} actualizada.")

    cursor.close()
    conexion.close()


# Interfaz gráfica
def hex_a_rgb(hex_color):
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

def crear_degradado_vertical(canvas, ancho, alto, color_inicio, color_fin):
    canvas.delete("degradado")  # Limpiar el canvas antes de dibujar un nuevo degradado

    rgb_inicio = hex_a_rgb(color_inicio)
    rgb_fin = hex_a_rgb(color_fin)

    pasos = alto // 2
    for i in range(pasos):
        y = alto - i - 1
        r = int(rgb_inicio[0] + (rgb_fin[0] - rgb_inicio[0]) * i / pasos)
        g = int(rgb_inicio[1] + (rgb_fin[1] - rgb_inicio[1]) * i / pasos)
        b = int(rgb_inicio[2] + (rgb_fin[2] - rgb_inicio[2]) * i / pasos)
        color = f"#{r:02x}{g:02x}{b:02x}"
        canvas.create_line(0, y, ancho, y, fill=color, tags="degradado")

    canvas.create_rectangle(0, 0, ancho, alto // 2, fill=color_fin, outline="", tags="degradado")


def gestionar_solicitudes(rol,volver_menu_callback):
    ventana = tk.Tk()
    ventana.title("Solicitudes de Pago")
    centrar_ventana(ventana, 1020, 600)

    canvas = tk.Canvas(ventana)
    canvas.pack(fill="both", expand=True)


    def actualizar_degradado(event):
        # Obtener las dimensiones del canvas
        ancho = canvas.winfo_width()
        alto = canvas.winfo_height()
        crear_degradado_vertical(canvas, ancho, alto, "#8B0000", "#FFFFFF")

    # Actualizar el fondo degradado al cambiar el tamaño de la ventana
    canvas.bind("<Configure>", actualizar_degradado)

    # Inicializar el degradado en el tamaño actual de la ventana
    ventana.after(100, lambda: actualizar_degradado(None))

    RUTA_LOGO = ruta_relativa("Plantillas/LogoATM.png")
    RUTA_LOGO2 = ruta_relativa("Plantillas/ISO-9001.jpeg")
    RUTA_LOGO3 = ruta_relativa("Plantillas/ISO-14001.jpeg")
    RUTA_LOGO4 = ruta_relativa("Plantillas/ISO-45001.jpeg")
    try:
        # LOGO ATM
        imagen = Image.open(RUTA_LOGO)
        imagen = imagen.resize((120, 160), Image.Resampling.LANCZOS)
        logo_img = ImageTk.PhotoImage(imagen)
        label_logo = tk.Label(canvas, image=logo_img, borderwidth=0)
        label_logo.image = logo_img
        label_logo.place(relx=0.05, rely=0.015)

        # ISO 9001
        iso1 = Image.open(RUTA_LOGO2).resize((60, 60), Image.Resampling.LANCZOS)
        iso1_img = ImageTk.PhotoImage(iso1)
        label_iso1 = tk.Label(canvas, image=iso1_img, borderwidth=0, bg="#ffffff")
        label_iso1.image = iso1_img
        label_iso1.place(relx=0.30, rely=0.020, anchor="ne")  # Esquina inferior derecha

        # ISO 14001
        iso2 = Image.open(RUTA_LOGO3).resize((60, 60), Image.Resampling.LANCZOS)
        iso2_img = ImageTk.PhotoImage(iso2)
        label_iso2 = tk.Label(canvas, image=iso2_img, borderwidth=0, bg="#ffffff")
        label_iso2.image = iso2_img
        label_iso2.place(relx=0.35, rely=0.13, anchor="ne")  # Al lado izquierdo del ISO 9001

        # ISO 45001
        iso3 = Image.open(RUTA_LOGO4).resize((60, 60), Image.Resampling.LANCZOS)
        iso3_img = ImageTk.PhotoImage(iso3)
        label_iso3 = tk.Label(canvas, image=iso3_img, borderwidth=0, bg="#ffffff")
        label_iso3.image = iso3_img
        label_iso3.place(relx=0.25, rely=0.13, anchor="ne")  # Al lado izquierdo del ISO 14001

    except Exception as e:
        print(f"⚠️ No se pudo cargar el logotipo: {e}")
        label_logo = tk.Label(canvas, text="LOGO", font=("Arial", 20, "bold"))


    # Buscar
    tk.Label(ventana, text="Buscar:", font=("Arial", 10, "bold"), bg="white").place(relx=0.05, rely=0.30)
    entry_busqueda = tk.Entry(ventana, width=50)
    entry_busqueda.place(relx=0.13, rely=0.30)

    def buscar(*args):  # Acepta *args por el trace
        termino = entry_busqueda.get().lower()
        for item in tree.get_children():
            tree.delete(item)

        conexion = conectar_bd()
        cursor = conexion.cursor()
        if termino:
            consulta = """
                SELECT id_autorizacion, fecha_solicitud, monto FROM autorizacionescompra
                WHERE LOWER(id_autorizacion) LIKE %s
                OR LOWER(fecha_solicitud) LIKE %s
                OR LOWER(monto) LIKE %s       
            """
            like_termino = f"%{termino}%"
            cursor.execute(consulta, (like_termino, like_termino, like_termino))

        else : 
            consulta= ("""
                SELECT id_autorizacion, fecha_solicitud, monto
                FROM autorizacionescompra
                WHERE id_autorizacion NOT IN (
                    SELECT id_autorizacion FROM SolicitudesPago
                )""")
            cursor.execute(consulta)
            
        resultados = cursor.fetchall()
        conexion.close()

        for row in resultados:
            tree.insert("", tk.END, values=row)

    entry_busqueda_var = tk.StringVar()
    entry_busqueda.config(textvariable=entry_busqueda_var)
    entry_busqueda_var.trace("w", buscar)  # Búsqueda automática al escribir

    tk.Label(ventana, text="Ingrese la informacion de la solicitud:", font=("Arial", 12, "bold"), bg="white").place(relx=0.60, rely=0.02)

    # Consecutivo
    tk.Label(ventana, text="Consecutivo:", font=("Arial", 10, "bold"), bg="white").place(relx=0.60, rely=0.1)
    entry_consecutivo = tk.Entry(ventana, width=30)
    entry_consecutivo.place(relx=0.70, rely=0.1)

    # Concepto
    tk.Label(ventana, text="Concepto:", font=("Arial", 10, "bold"), bg="white").place(relx=0.60, rely=0.16)
    entry_concepto = tk.Text(ventana, wrap="word", font=("Arial", 10))
    entry_concepto.place(relx=0.70, rely=0.14, relwidth=0.23, relheight=0.06)

    # Referencia de Pago
    tk.Label(ventana, text="Referencia de Pago:", font=("Arial", 10, "bold"), bg="white").place(relx=0.60, rely=0.22)
    entry_referencia = tk.Entry(ventana, width=30)
    entry_referencia.place(relx=0.75, rely=0.22)

    #Numero de Factura
    tk.Label(ventana, text="Numero de Factura:", font=("Arial", 10, "bold"), bg="white").place(relx=0.60, rely=0.28)
    entry_factura = tk.Entry(ventana,width=15)
    entry_factura.place(relx=0.75, rely=0.28)
    tk.Button(ventana, text="Cargar Factura/Cotizacion", command=lambda: seleccionar_factura(entry_factura)).place(relx=0.85, rely=0.28)
    tk.Button(ventana, text= "Cambiar Factura", command=lambda: eliminar_factura(entry_factura)).place(relx=0.85, rely=0.33)


    #Aplicacion del estilo a la tabla
    style = ttk.Style()
    style.theme_use("alt")
    style.configure("Treeview.Heading", font=("Arial", 10, "bold"), foreground="white", background="#990000")
    style.map("Treeview.Heading", background=[("!active", "#990000"), ("active", "#990000"), ("pressed", "#990000")],
              foreground=[("!active", "white"), ("active", "white"), ("pressed", "white")])

    # Treeview (tabla)
    tree = ttk.Treeview(ventana, columns=("ID", "Fecha", "Monto", "Limite de Pago"), show="headings")
    for col in ("ID", "Fecha", "Monto", "Limite de Pago"):
        tree.heading(col, text=col)
        tree.column(col, anchor="center")

    # Scrollbar
    scrollbar = ttk.Scrollbar(ventana, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)

    tree.place(relx=0.05, rely=0.38, relwidth=0.88, relheight=0.50)
    scrollbar.place(relx=0.93, rely=0.38, relheight=0.50)

    # Ventana de solicitudes guardadas
    def Solicitudes(parent=None):
        ventana_solicitudes = tk.Toplevel(parent) if parent else tk.Toplevel()
        ventana_solicitudes.title("Solicitudes Guardadas")
        centrar_ventana(ventana_solicitudes, 1100, 600)

        #Aplicacion del estilo a la tabla
        style = ttk.Style()
        style.theme_use("alt")
        style.configure("Treeview.Heading", font=("Arial", 10, "bold"), foreground="white", background="#990000")
        style.map("Treeview.Heading", background=[("!active", "#990000"), ("active", "#990000"), ("pressed", "#990000")],
                foreground=[("!active", "white"), ("active", "white"), ("pressed", "white")])

        
        frame_tabla = tk.Frame(ventana_solicitudes)
        frame_tabla.place(relx=0.05, rely=0.05, relwidth=0.9, relheight=0.75)  # ⬅️ Ajusta tamaño aquí

        tree_local = ttk.Treeview(frame_tabla, columns=("ID", "Fecha", "Importe", "Limite de Pago", "Estado"), show="headings")
        for col in ("ID", "Fecha", "Importe", "Limite de Pago","Estado"):
            tree_local.heading(col, text=col)
            tree_local.column(col, width=60, anchor="center")

        tree_local.place(relx=0, rely=0, relwidth=0.97, relheight=1)  # ⬅️ Ajusta si quieres más separación

        scrollbar_local = ttk.Scrollbar(frame_tabla, orient="vertical", command=tree_local.yview)
        tree_local.configure(yscrollcommand=scrollbar_local.set)
        scrollbar_local.place(relx=0.97, rely=0, relwidth=0.03, relheight=1)  # ⬅️ Posición vertical a la derecha

        tk.Label(ventana_solicitudes, text="Buscar ID:", font=("Arial", 10, "bold")).place(relx=0.1, rely=0.85)
        entry_busqueda = ttk.Entry(ventana_solicitudes, width=20)
        entry_busqueda.place(relx=0.2, rely=0.85)

        def buscar_solicitud():
            id_buscar = entry_busqueda.get().strip()
            for item in tree_local.get_children():
                tree_local.selection_remove(item)
                valores = tree_local.item(item, "values")
                if valores and str(valores[0]) == id_buscar:
                    tree_local.see(item)
                    tree_local.selection_add(item)
                    break
            else:
                ventana_solicitudes.lift()
                ventana_solicitudes.focus_force()
                messagebox.showinfo("No encontrado", f"No se encontró la solicitud con ID {id_buscar}")

        tk.Button(ventana_solicitudes, text="Buscar", command=buscar_solicitud,
                  font=("Arial", 10, "bold"), bg="#004080", fg="white").place(relx=0.35, rely=0.845)

        tk.Button(ventana_solicitudes, text="Salir", command=ventana_solicitudes.destroy,
                bg="red", fg="white", font=("Arial", 10, "bold")).place(relx=0.1, rely=0.92, relwidth=0.08, relheight=0.04)
        
        tk.Button(ventana_solicitudes, text="Marcar como Pagado", font=("Arial", 10, "bold"),
          command=lambda: marcar_como_pagado(tree_local, usuario_actual),
          bg="#006400", fg="white").place(relx=0.6, rely=0.92, relwidth=0.2, relheight=0.06)

        cargar_solicitudes(tree_local)

    # Botones
    tk.Button(ventana, text="Guardar y Generar Docs",
              command=lambda: generar_excel_desde_seleccion(tree, entry_consecutivo, entry_concepto, entry_referencia, entry_factura), font=("Arial", 10, "bold")
             ).place(relx=0.35, rely=0.91, relwidth=0.18, relheight=0.05)

    tk.Button(ventana, text="Solicitudes Guardadas", command=Solicitudes, font=("Arial", 10, "bold")).place(relx=0.75, rely=0.91, relwidth=0.15, relheight=0.05)
    tk.Button(ventana, text="Salir", command= lambda: salir(volver_menu_callback, ventana), bg="red", fg="white", font=("Arial", 10, "bold")).place(relx=0.1, rely=0.91, relwidth=0.095, relheight=0.05)
    tk.Button(ventana, text="Reporte de Costos", command=lambda: costos_contrato(), font=("Arial", 10, "bold")
              ).place(relx=0.55, rely=0.91, relwidth=0.15, relheight=0.05)

    cargar_autorizaciones(tree)
    ventana.mainloop()