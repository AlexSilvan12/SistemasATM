import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageTk
from openpyxl import load_workbook
from datetime import datetime
from database import conectar_bd
from gastos_contrato import crear_degradado_vertical
from utils import centrar_ventana, ruta_relativa

def exportar_solicitudes(mes, año):
   

    # Conexión a la base de datos
    conn = conectar_bd()
    cursor = conn.cursor()

    # Consulta SQL filtrada por mes y año
    cursor.execute("""
        SELECT 
            sp.fecha_solicitud, 
            sp.id_solicitud, 
            p.nombre, 
            sp.concepto, 
            CASE WHEN sp.moneda = 'MXN' THEN sp.importe ELSE NULL END AS monto_mxn,
            CASE WHEN sp.moneda = 'USD' THEN sp.importe ELSE NULL END AS monto_usd,
            sp.id_autorizacion, 
            sp.fecha_pago, 
            sp.estado, 
            sp.fecha_limite_pago
        FROM solicitudespago sp
        INNER JOIN proveedores p ON sp.id_proveedor = p.id_proveedor
        WHERE MONTH(sp.fecha_solicitud) = %s AND YEAR(sp.fecha_solicitud) = %s
        ORDER BY sp.fecha_solicitud
    """, (mes, año))
    datos = cursor.fetchall()

    if not datos:
        messagebox.showinfo("Sin datos", f"No se encontraron solicitudes para {mes}/{año}.")
        cursor.close()
        conn.close()
        return

    # Cargar plantilla
    ruta_plantilla = ruta_relativa("Plantillas/Control de solicitudes.xlsx")
    wb = load_workbook(ruta_plantilla)
    ws = wb.active

    # Insertar datos a partir de fila 9
    fila = 9
    for row in datos:
        fecha_solicitud, id_solicitud, nombre, concepto, monto_mxn, monto_usd, id_autorizacion, fecha_pago, estado, fecha_limite_pago = row
        ws[f"B{fila}"] = fecha_solicitud
        ws[f"C{fila}"] = id_solicitud
        ws[f"D{fila}"] = nombre
        ws[f"E{fila}"] = concepto
        ws[f"F{fila}"] = monto_mxn
        ws[f"G{fila}"] = monto_usd
        ws[f"H{fila}"] = id_autorizacion
        ws[f"M{fila}"] = fecha_pago
        ws[f"L{fila}"] = estado
        ws[f"N{fila}"] = fecha_limite_pago
        fila += 1

    # Diálogo para guardar archivo
    ruta_guardado = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Archivos Excel", "*.xlsx")],
        title="Guardar reporte como",
        initialfile=f"Control_de_solicitudes_{año}_{mes:02d}.xlsx"
    )

    if ruta_guardado:
        wb.save(ruta_guardado)
        messagebox.showinfo("Éxito", f"Archivo guardado en:\n{ruta_guardado}")

    # Cerrar conexión
    cursor.close()
    conn.close()

# --- Interfaz gráfica ---
def Control_Solicitudes(usuario_actual, ventana_padre):
    if usuario_actual["rol"] != "Contador" and usuario_actual["rol"] !="Administrador":
        messagebox.showwarning("Acceso denegado", "No tiene permisos para realizar esta acción.", parent=ventana_padre)
        return
    
    ventana = tk.Toplevel()
    ventana.title("Filtrar Solicitudes por Mes y Año")
    centrar_ventana(ventana, 600, 300)

    canvas = tk.Canvas(ventana)
    canvas.pack(fill="both", expand=True)

    def actualizar_degradado(event):
        ancho = canvas.winfo_width()
        alto = canvas.winfo_height()
        crear_degradado_vertical(canvas, ancho, alto, "#8B0000", "#FFFFFF")

    canvas.bind("<Configure>", actualizar_degradado)
    ventana.after(100, lambda: actualizar_degradado(None))

    RUTA_LOGO = ruta_relativa("Plantillas/LogoATM.png")
    RUTA_LOGO2 = ruta_relativa("Plantillas/ISO-9001.jpeg")
    RUTA_LOGO3 = ruta_relativa("Plantillas/ISO-14001.jpeg")
    RUTA_LOGO4 = ruta_relativa("Plantillas/ISO-45001.jpeg")

    try:
        imagen = Image.open(RUTA_LOGO).resize((70, 90), Image.Resampling.LANCZOS)
        logo_img = ImageTk.PhotoImage(imagen)
        label_logo = tk.Label(canvas, image=logo_img, borderwidth=0)
        label_logo.image = logo_img
        label_logo.place(relx=0.02, rely=0.01)

        iso1_img = ImageTk.PhotoImage(Image.open(RUTA_LOGO2).resize((40, 40), Image.Resampling.LANCZOS))
        label_iso1 = tk.Label(canvas, image=iso1_img, borderwidth=0, bg="#ffffff")
        label_iso1.image = iso1_img
        label_iso1.place(relx=0.90, rely=0.05, anchor="ne")

        iso2_img = ImageTk.PhotoImage(Image.open(RUTA_LOGO3).resize((40, 40), Image.Resampling.LANCZOS))
        label_iso2 = tk.Label(canvas, image=iso2_img, borderwidth=0, bg="#ffffff")
        label_iso2.image = iso2_img
        label_iso2.place(relx=0.95, rely=0.18, anchor="ne")

        iso3_img = ImageTk.PhotoImage(Image.open(RUTA_LOGO4).resize((40, 40), Image.Resampling.LANCZOS))
        label_iso3 = tk.Label(canvas, image=iso3_img, borderwidth=0, bg="#ffffff")
        label_iso3.image = iso3_img
        label_iso3.place(relx=0.85, rely=0.18, anchor="ne")

    except Exception as e:
        print(f"⚠️ No se pudo cargar el logotipo: {e}")
        tk.Label(canvas, text="LOGO", font=("Arial", 20, "bold")).place(relx=0.07, rely=0.01)

    label_titulo = tk.Label(canvas, text="ATM | Control de Solicitudes",
                             font=("Arial", 17, "bold"), bg="white")
    label_titulo.place(relx=0.5, rely=0.02, anchor="n")

    # Mes
    tk.Label(ventana, text="Mes:", font=("Arial", 10,"bold"), bg="white").place(relx=0.40, rely=0.45, anchor="e")
    combo_mes = ttk.Combobox(ventana, values=[
        "1 - Enero", "2 - Febrero", "3 - Marzo", "4 - Abril",
        "5 - Mayo", "6 - Junio", "7 - Julio", "8 - Agosto",
        "9 - Septiembre", "10 - Octubre", "11 - Noviembre", "12 - Diciembre"
    ], font=("Arial", 10,"bold"))
    combo_mes.place(relx=0.45, rely=0.45, anchor="w")

    # Año
    tk.Label(ventana, text="Año:", font=("Arial", 10,"bold"), bg="#FFF5F5").place(relx=0.40, rely=0.55, anchor="e")
    combo_año = ttk.Combobox(ventana, values=list(range(2020, datetime.now().year + 1)), font=("Arial", 10,"bold"))
    combo_año.place(relx=0.45, rely=0.55, anchor="w")

    def ejecutar():
        if not combo_mes.get() or not combo_año.get():
            messagebox.showwarning("Datos incompletos", "Selecciona mes y año.")
            return
        mes_num = int(combo_mes.get().split(" - ")[0])
        año_num = int(combo_año.get())
        exportar_solicitudes(mes_num, año_num)

    tk.Button(ventana, text="Generar Excel", font=("Arial", 10), bg= "#990000", fg= "white", command=ejecutar).place(relx=0.5, rely=0.75, anchor="center")
